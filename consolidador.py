#!/usr/bin/env python3
"""
consolidador.py
───────────────
Extrae datos de PDFs (GBM / Prestadero) y los integra en el archivo
maestro del mes anterior para generar el nuevo Estado de Cuenta Consolidado.

Flujo:
  1. Lee PDFs de PDFs_Origen/
  2. Lee el archivo maestro de 00_Maestro_Anterior/
  3. Actualiza cada hoja del maestro con los datos nuevos
  4. Guarda el resultado en Resultados_Excel/
"""

import pdfplumber
import os
import re
import shutil
from copy import copy
from collections import defaultdict
from openpyxl import load_workbook

# ════════════════════════════════════════════════════════════
# RUTAS (relativas al directorio del script)
# ════════════════════════════════════════════════════════════
BASE = os.path.dirname(os.path.abspath(__file__))
CARPETA_PDFS = os.path.join(BASE, "PDFs_Origen")
CARPETA_MAESTRO = os.path.join(BASE, "00_Maestro_Anterior")
CARPETA_SALIDA = os.path.join(BASE, "Estados de cuenta del mes")

os.makedirs(CARPETA_SALIDA, exist_ok=True)

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
MESES_INV = {v: k for k, v in MESES.items()}


# ════════════════════════════════════════════════════════════
# FUNCIONES DE EXTRACCIÓN  (idénticas a extractor_gbm.py)
# ════════════════════════════════════════════════════════════
def extraer_numeros(texto):
    nums = re.findall(r"[\d,]+\.\d+", texto)
    return [float(n.replace(",", "")) for n in nums]


def extraer_todos_numeros(texto):
    nums = re.findall(r"[\d,]+\.?\d*", texto)
    return [float(n.replace(",", "")) for n in nums if n]


def extraer_numero_despues_de(texto, clave):
    idx = texto.find(clave)
    if idx == -1:
        return None
    sub = texto[idx + len(clave):]
    nums = extraer_numeros(sub)
    return nums[0] if nums else None


def extraer_nombre_cliente(pdf, plataforma):
    texto = pdf.pages[0].extract_text() or ""
    lineas = texto.split("\n")
    if plataforma == "Prestadero":
        for l in lineas:
            if "Periodo:" in l and "Estado de Cuenta" not in l:
                return re.split(r"\s+Periodo:", l)[0].strip().upper()
    else:
        for l in lineas:
            if "Contrato:" in l:
                p = re.split(r"\s+Contrato:", l)[0].strip()
                p = p.replace("PUBLICO EN GENERAL - ", "")
                return p.upper()
    return "DESCONOCIDO"


def detectar_plataforma(texto):
    return "Prestadero" if ("Prestadero" in texto or "PRESTADERO" in texto) else "GBM"


def es_smart_cash(texto):
    for l in texto.split("\n"):
        if "RENTA VARIABLE" in l and "VALORES EN CORTO" not in l:
            nums = extraer_numeros(l)
            if len(nums) >= 2 and nums[1] > 0:
                return False
    return True


def extraer_saldo_anterior(lineas):
    for l in lineas:
        if "VALOR DEL PORTAFOLIO" in l and "TOTAL" not in l:
            nums = extraer_numeros(l)
            if nums:
                return nums[0]
    return 0.0


def extraer_portafolio_gbm(pdf):
    portafolio = []
    en_desglose = en_acciones = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu:
                en_desglose = True; continue
            if not en_desglose:
                continue
            if lu == "ACCIONES":
                en_acciones = True; continue
            if en_acciones and ("EMISORA" in lu or "MES ANTERIOR" in lu or "EN PR" in lu):
                continue
            if en_acciones and lu.startswith("TOTAL"):
                en_acciones = False; continue
            if lu in ("DESGLOSE DE MOVIMIENTOS", "RENDIMIENTO DEL PORTAFOLIO", "EFECTIVO"):
                en_desglose = en_acciones = False; continue
            if not en_acciones:
                continue
            m = re.match(r"^([A-Z]+(?:\s+\d+)?)\s+", ls)
            if m:
                try:
                    emisora = m.group(1).strip()
                    nums = extraer_todos_numeros(ls[m.end():])
                    if len(nums) >= 8:
                        portafolio.append({
                            "Emisora": emisora,
                            "Títulos Mes Anterior": int(nums[0]),
                            "Títulos Mes Actual": int(nums[1]),
                            "Costo Total": nums[4],
                            "Precio Mercado Mes Anterior": nums[5],
                            "Precio Mercado Mes Actual": nums[6],
                            "Valor a Mercado": nums[7],
                        })
                except (ValueError, IndexError):
                    continue
    return portafolio


def extraer_deuda_gbm(pdf):
    deuda = []
    en_desglose = en_deuda = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu:
                en_desglose = True; continue
            if not en_desglose:
                continue
            if "DEUDA EN REPORTO" in lu and "TOTAL" not in lu:
                en_deuda = True; continue
            if en_deuda and ("EMISORA" in lu or "ANTERIOR" in lu):
                continue
            if en_deuda and lu.startswith("TOTAL"):
                en_deuda = False; continue
            if lu in ("RENTA VARIABLE", "EFECTIVO", "DESGLOSE DE MOVIMIENTOS", "RENDIMIENTO DEL PORTAFOLIO"):
                en_desglose = en_deuda = False; continue
            if not en_deuda:
                continue
            m = re.match(r"^([A-Z]+\s+\d+)\s+", ls)
            if m:
                try:
                    emisora = m.group(1).strip()
                    nums = extraer_todos_numeros(ls[m.end():])
                    if len(nums) >= 8:
                        deuda.append({
                            "Emisora": emisora,
                            "Títulos Mes Anterior": int(nums[0]),
                            "Títulos Mes Actual": int(nums[1]),
                            "Tasa": nums[2],
                            "Valor del Reporto": nums[7],
                            "% Cartera": nums[9] if len(nums) >= 10 else 0.0,
                        })
                except (ValueError, IndexError):
                    continue
    return deuda


def extraer_movimientos_acciones(pdf):
    movimientos = []
    en_mov = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            lu = l.strip().upper()
            if "DESGLOSE DE MOVIMIENTOS" in lu:
                en_mov = True; continue
            if en_mov and lu in ("RENDIMIENTO DEL PORTAFOLIO", "COMPOSICIÓN FISCAL INFORMATIVA"):
                en_mov = False; continue
            if not en_mov:
                continue
            if "Compra de Acciones" not in l and "Venta de Acciones" not in l:
                continue
            try:
                ls = l.strip()
                if "Compra de Acciones" in l:
                    op = "Compra"
                    resto = l[l.find("Compra de Acciones.") + len("Compra de Acciones."):].strip()
                else:
                    op = "Venta"
                    resto = l[l.find("Venta de Acciones.") + len("Venta de Acciones."):].strip()
                fm = re.match(r"(\d{2}/\d{2})", ls)
                fecha = fm.group(1) if fm else ""
                em = re.match(r"^([A-Z]+(?:\s+\d+)?)\s+", resto)
                if em:
                    emisora = em.group(1).strip()
                    nums = extraer_todos_numeros(resto[em.end():])
                    movimientos.append({
                        "Fecha": fecha,
                        "Operación": op,
                        "Emisora": emisora,
                        "Títulos": int(nums[0]) if nums else 0,
                        "Precio Unitario": nums[1] if len(nums) >= 2 else 0,
                        "Comisión": nums[2] if len(nums) >= 3 else 0,
                        "Neto": nums[5] if len(nums) >= 6 else 0,
                    })
            except Exception:
                continue
    return movimientos


def extraer_periodo_pdf(pdf, plataforma):
    """Extrae mes, año y cadena de periodo del PDF."""
    texto = pdf.pages[0].extract_text() or ""
    if plataforma == "Prestadero":
        m = re.search(r"(\d{4}).(\d{2}).(\d{2})\s+al\s+(\d{4}).(\d{2}).(\d{2})", texto)
        if m:
            anio, mes = int(m.group(4)), int(m.group(5))
            dia_fin = int(m.group(6))
            nombre_mes = MESES_INV.get(mes, str(mes))
            return {"mes": mes, "anio": anio, "mes_nombre": nombre_mes,
                    "periodo": f"01-{dia_fin} {nombre_mes[:3]} DE {anio}"}
    else:
        m = re.search(r"DEL\s+(\d+)\s+AL\s+(\d+)\s+DE\s+(\w+)\s+DE\s+(\d{4})", texto, re.I)
        if m:
            dia_ini, dia_fin = int(m.group(1)), int(m.group(2))
            nombre_mes = m.group(3).upper()
            anio = int(m.group(4))
            mes = MESES.get(nombre_mes, 0)
            return {"mes": mes, "anio": anio, "mes_nombre": nombre_mes,
                    "periodo": f"{dia_ini:02d}-{dia_fin} {nombre_mes[:3]} DE {anio}"}
    return None


# ════════════════════════════════════════════════════════════
# PASO 1:  EXTRAER DATOS DE TODOS LOS PDFs
# ════════════════════════════════════════════════════════════
def extraer_todos_los_pdfs():
    """Procesa cada PDF y agrupa la información por nombre de cliente."""
    clientes = defaultdict(lambda: {
        "gbm": None, "smart_cash": None, "prestadero": None, "periodo": None
    })

    for archivo in sorted(os.listdir(CARPETA_PDFS)):
        if not archivo.lower().endswith(".pdf"):
            continue
        ruta = os.path.join(CARPETA_PDFS, archivo)
        print(f"  📄 Leyendo: {archivo}")

        try:
            with pdfplumber.open(ruta) as pdf:
                texto_p1 = pdf.pages[0].extract_text() or ""
                texto_completo = texto_p1
                for p in pdf.pages[1:]:
                    t = p.extract_text()
                    if t:
                        texto_completo += "\n" + t

                plataforma = detectar_plataforma(texto_completo)
                nombre = extraer_nombre_cliente(pdf, plataforma)
                lineas_p1 = texto_p1.split("\n")
                periodo = extraer_periodo_pdf(pdf, plataforma)

                print(f"     → Cliente: {nombre} | {plataforma}")

                if periodo and (clientes[nombre]["periodo"] is None
                                or periodo["mes"] > clientes[nombre]["periodo"]["mes"]):
                    clientes[nombre]["periodo"] = periodo

                # ─── Prestadero ───
                if plataforma == "Prestadero":
                    abonos = retiros = interes = valor = 0.0
                    for l in lineas_p1:
                        try:
                            if "Abonos:" in l and "Cuenta Abonos:" not in l:
                                v = extraer_numero_despues_de(l, "Abonos:")
                                if v is not None: abonos = v
                            if "Valor de la Cuenta:" in l:
                                v = extraer_numero_despues_de(l, "Valor de la Cuenta:")
                                if v is not None: valor = v
                            if "Interés Recibido" in l or "Interes Recibido" in l:
                                ns = extraer_numeros(l)
                                if ns: interes = ns[0]
                            if "Retiros:" in l and "Detalle" not in l:
                                v = extraer_numero_despues_de(l, "Retiros:")
                                if v is not None: retiros = v
                        except Exception:
                            continue
                    clientes[nombre]["prestadero"] = {
                        "abonos": abonos, "retiros": retiros,
                        "interes": interes, "valor": valor,
                    }

                # ─── GBM ───
                else:
                    entradas = salidas = valor_total = saldo_ant = 0.0
                    saldo_ant = extraer_saldo_anterior(lineas_p1)
                    for l in lineas_p1:
                        try:
                            if "ENTRADAS DE EFECTIVO" in l:
                                ns = extraer_numeros(l)
                                if ns: entradas = ns[-1]
                            elif "SALIDAS DE EFECTIVO" in l:
                                ns = extraer_numeros(l)
                                if ns: salidas = ns[-1]
                            elif "VALOR DEL PORTAFOLIO" in l and "TOTAL" not in l:
                                ns = extraer_numeros(l)
                                if len(ns) >= 2: valor_total = ns[1]
                                elif ns: valor_total = ns[0]
                        except Exception:
                            continue

                    smart = es_smart_cash(texto_p1)
                    portafolio = deuda = movimientos = []
                    try: deuda = extraer_deuda_gbm(pdf)
                    except Exception: pass
                    if not smart:
                        try: portafolio = extraer_portafolio_gbm(pdf)
                        except Exception: pass
                        try: movimientos = extraer_movimientos_acciones(pdf)
                        except Exception: pass

                    tipo = "smart_cash" if smart else "gbm"
                    print(f"     → Tipo: {'Smart Cash' if smart else 'GBM Regular'}")

                    clientes[nombre][tipo] = {
                        "entradas": entradas, "salidas": salidas,
                        "valor_total": valor_total, "saldo_anterior": saldo_ant,
                        "portafolio": portafolio, "deuda": deuda,
                        "movimientos": movimientos,
                    }
        except Exception as e:
            print(f"     ❌ Error: {e}")

    return clientes


# ════════════════════════════════════════════════════════════
# FUNCIONES DE CONSOLIDACIÓN
# ════════════════════════════════════════════════════════════
def normalizar(nombre):
    """Normaliza nombre de instrumento para comparación."""
    if not nombre or nombre == "-":
        return ""
    n = str(nombre).upper().strip()
    n = re.sub(r"\n", " ", n)
    n = re.sub(r"\s+", " ", n)
    return n


# Tabla de alias para hacer match entre PDF y maestro
ALIASES = {
    "FIBRAPL 14": ["FIBRA PL 14", "FIBRA PL14", "FIBRAPL14"],
    "FIHO 12":    ["FIHO12"],
    "FMTY 14":    ["FMTY14"],
    "FUNO 11":    ["FUNO11"],
    "FIBRAMQ 12": ["FIBRAMQ12"],
    "DAHANOS 13": ["DANHOS 13", "DANHOS13", "DAHANOS13"],
}

# Construir mapa bidireccional de alias
_ALIAS_MAP = {}
for canonical, alts in ALIASES.items():
    group = {normalizar(canonical)} | {normalizar(a) for a in alts}
    for name in group:
        _ALIAS_MAP[name] = group


def instrumentos_coinciden(nombre_pdf, nombre_master):
    """¿Dos nombres de instrumento se refieren a lo mismo?"""
    np_ = normalizar(nombre_pdf)
    nm_ = normalizar(nombre_master)
    if np_ == nm_:
        return True
    grupo = _ALIAS_MAP.get(np_, {np_})
    return nm_ in grupo


def encontrar_fila(ws, texto_buscar, col=1, rango=(1, 50)):
    """Busca una fila que contenga cierto texto en la columna dada."""
    for r in range(rango[0], rango[1]):
        v = ws.cell(r, col).value
        if v and texto_buscar in str(v).upper():
            return r
    return None


def valor_numerico(v, default=0.0):
    """Convierte un valor de celda a float, o devuelve default."""
    if isinstance(v, (int, float)):
        return float(v)
    return default


def actualizar_celda(ws, row, col, value, forzar=False):
    """Escribe en una celda respetando merged ranges y fórmulas.
    Si la celda contiene una fórmula (empieza con '='), NO la sobreescribe
    a menos que forzar=True.
    """
    # Resolver merged range → celda superior izquierda
    target_row, target_col = row, col
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            target_row, target_col = rng.min_row, rng.min_col
            break

    celda = ws.cell(target_row, target_col)
    # Preservar fórmulas existentes
    if not forzar and isinstance(celda.value, str) and celda.value.startswith("="):
        return
    celda.value = value


def copiar_formato_fila(ws, fila_origen, fila_destino):
    """Copia formato (font, fill, border, number_format, alignment) de una
    fila a otra, columnas A-O (1-15)."""
    for col in range(1, 16):
        src = ws.cell(fila_origen, col)
        dst = ws.cell(fila_destino, col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def insertar_instrumento(ws, fila_totales, datos, fila_ref, periodo):
    """Inserta una fila para un instrumento nuevo justo antes de TOTALES.

    Retorna la nueva posición de TOTALES (fila_totales + 1).
    """
    ws.insert_rows(fila_totales)
    nueva_fila = fila_totales          # la fila recién insertada

    copiar_formato_fila(ws, fila_ref, nueva_fila)

    emisora = datos["emisora"]
    valor = datos["valor_a_mercado"]
    compra = datos["compra_neto"]
    venta = datos["venta_neto"]

    b = compra if compra > 0 else valor   # inversión original
    c = valor
    e = c - b
    f = (e / b) if b > 0 else 0.0
    g = c - 0 + venta - compra            # old_c=0 para instrumento nuevo
    h = (g / b) if b > 0 else 0.0

    ws.cell(nueva_fila, 1).value = emisora            # A
    ws.cell(nueva_fila, 2).value = round(b, 2)        # B
    ws.cell(nueva_fila, 3).value = round(c, 2)        # C
    if periodo:                                         # D
        ws.cell(nueva_fila, 4).value = f"{periodo['mes_nombre'].lower()} {periodo['anio']}"
    ws.cell(nueva_fila, 5).value = round(e, 2)        # E
    ws.cell(nueva_fila, 6).value = round(f, 10)       # F
    ws.cell(nueva_fila, 7).value = round(g, 2)        # G
    ws.cell(nueva_fila, 8).value = round(h, 10)       # H
    ws.cell(nueva_fila, 9).value = "RENTA VARIABLE"   # I
    ws.cell(nueva_fila, 10).value = round(venta, 2)   # J
    ws.cell(nueva_fila, 11).value = round(compra, 2)  # K
    ws.cell(nueva_fila, 14).value = round(c, 2)       # N
    ws.cell(nueva_fila, 15).value = "GBM"             # O

    return fila_totales + 1


def expandir_formulas_totales(ws, fila_totales):
    """Expande los rangos =SUM(…) en la fila TOTALES para que incluyan
    todas las filas de instrumentos hasta fila_totales - 1."""
    patron = re.compile(r"(SUM\([A-Z]+)(\d+)(:[A-Z]+)(\d+)(\))")
    nueva_fin = fila_totales - 1
    for col in range(2, 16):
        celda = ws.cell(fila_totales, col)
        val = celda.value
        if not isinstance(val, str) or not val.startswith("="):
            continue
        nueva_formula = patron.sub(
            lambda m: f"{m.group(1)}{m.group(2)}{m.group(3)}{nueva_fin}{m.group(5)}",
            val,
        )
        if nueva_formula != val:
            celda.value = nueva_formula


def leer_instrumentos_master(ws, fila_header, fila_totales):
    """Lee todas las filas de instrumentos del maestro."""
    instrumentos = []
    r = fila_header + 1
    while r < fila_totales:
        nombre = ws.cell(r, 1).value
        if nombre and str(nombre).strip() and str(nombre).strip() != "-":
            # Determinar si la fila está merged verticalmente
            fila_fin = r
            for rng in ws.merged_cells.ranges:
                if rng.min_row == r and rng.min_col == 1:
                    fila_fin = rng.max_row
                    break

            instrumentos.append({
                "fila": r,
                "fila_fin": fila_fin,
                "nombre": str(nombre).strip(),
                "B": ws.cell(r, 2).value,   # Saldo Inicial
                "C": ws.cell(r, 3).value,   # Saldo Total
                "D": ws.cell(r, 4).value,   # Fecha Inicial
                "E": ws.cell(r, 5).value,   # Ganancia Histórica
                "F": ws.cell(r, 6).value,   # Ganancia Histórica %
                "G": ws.cell(r, 7).value,   # Ganancia / Pérdida Mes
                "H": ws.cell(r, 8).value,   # Ganancia / Pérdida Mes %
                "I": ws.cell(r, 9).value,   # Clasificación
                "J": ws.cell(r, 10).value,  # Retiros
                "K": ws.cell(r, 11).value,  # Depósitos
                "L": ws.cell(r, 12).value,  # % Cartera
                "M": ws.cell(r, 13).value,  # Liquidez
                "N": ws.cell(r, 14).value,  # Total
                "O": ws.cell(r, 15).value,  # Proveedor
            })
            r = fila_fin + 1
        else:
            r += 1
    return instrumentos


# ════════════════════════════════════════════════════════════
# PASO 2:  ACTUALIZAR UNA HOJA DEL MAESTRO
# ════════════════════════════════════════════════════════════
def _mejor_match_deuda(old_c, fuentes_disponibles):
    """Busca la fuente cuyo valor sea más cercano al C anterior.
    Utiliza matching por proximidad sin límite estricto de tolerancia;
    simplemente elige la fuente más cercana.
    Retorna (key, fuente_dict) o (None, None).
    """
    if not fuentes_disponibles:
        return None, None

    if old_c <= 0:
        # Fila nueva o sin C previo: tomar la fuente más pequeña
        mejor_key = min(fuentes_disponibles,
                        key=lambda k: fuentes_disponibles[k]["valor"])
        return mejor_key, fuentes_disponibles[mejor_key]

    mejor_key = None
    mejor_diff = float("inf")
    for key, fuente in fuentes_disponibles.items():
        diff = abs(fuente["valor"] - old_c)
        if diff < mejor_diff:
            mejor_diff = diff
            mejor_key = key

    if mejor_key is not None:
        return mejor_key, fuentes_disponibles[mejor_key]

    return None, None


def actualizar_hoja(ws, datos, nombre_hoja):
    """Actualiza una hoja del maestro con los datos extraídos de PDFs.

    REGLAS CLAVE:
    ─ B (Saldo Inicial) para fibras/acciones = B_prev + K - J (se
      actualiza con compras y ventas). DEUDA y EFECTIVO: B no cambia.
    ─ Solo se tocan filas con coincidencia GBM/Prestadero. Externos
      (CAPITAL, Metales, GLD, SLV, NFLX, etc.) se dejan intactos.
    ─ Instrumentos nuevos se insertan antes de EFECTIVO GBM.
    ─ Secciones debajo de TOTALES se eliminan.
    ─ EFECTIVO GBM C = gbm_total - portafolio - deuda_gbm_mapeada.
    """

    gbm = datos.get("gbm")
    smart_cash = datos.get("smart_cash")
    prestadero = datos.get("prestadero")
    periodo = datos.get("periodo")

    # ── Localizar estructura ──
    fila_header = encontrar_fila(ws, "INSTRUMENTO") or 23
    fila_totales = encontrar_fila(ws, "TOTALES", rango=(fila_header, fila_header + 40))
    if not fila_totales:
        print(f"    ⚠️  No se encontró fila TOTALES en '{nombre_hoja}' — se omite.")
        return

    instrumentos = leer_instrumentos_master(ws, fila_header, fila_totales)
    print(f"    📊 Instrumentos encontrados: {len(instrumentos)}")

    # ── Preparar lookup de datos PDF ──
    pdf_port = {}
    compras_map = defaultdict(float)
    ventas_map = defaultdict(float)

    if gbm:
        for item in gbm.get("portafolio", []):
            key = normalizar(item["Emisora"])
            pdf_port[key] = item["Valor a Mercado"]

        for mov in gbm.get("movimientos", []):
            key = normalizar(mov["Emisora"])
            if mov["Operación"] == "Compra":
                compras_map[key] += mov["Neto"]
            else:
                ventas_map[key] += mov["Neto"]

    # Deuda GBM total
    deuda_gbm_total = 0.0
    if gbm:
        deuda_gbm_total = sum(d["Valor del Reporto"] for d in gbm.get("deuda", []))

    # Deuda Smart Cash total
    deuda_sc_total = 0.0
    sc_entradas = sc_salidas = 0.0
    if smart_cash:
        deuda_sc_total = sum(d["Valor del Reporto"] for d in smart_cash.get("deuda", []))
        sc_entradas = smart_cash.get("entradas", 0)
        sc_salidas = smart_cash.get("salidas", 0)

    # ── Pool de fuentes DEUDA disponibles (matching por proximidad) ──
    fuentes_deuda = {}
    if prestadero:
        fuentes_deuda["prestadero"] = {
            "valor": prestadero["valor"],
            "retiros": prestadero["retiros"],
            "depositos": prestadero["abonos"],
            "interes": prestadero["interes"],
            "tipo": "prestadero",
        }
    if smart_cash and deuda_sc_total > 0:
        fuentes_deuda["smart_cash"] = {
            "valor": deuda_sc_total,
            "retiros": sc_salidas,
            "depositos": sc_entradas,
            "interes": 0.0,
            "tipo": "smart_cash",
        }
    if gbm and deuda_gbm_total > 0:
        fuentes_deuda["gbm_deuda"] = {
            "valor": deuda_gbm_total,
            "retiros": 0.0,
            "depositos": 0.0,
            "interes": 0.0,
            "tipo": "gbm_deuda",
        }

    # ── Iterar instrumentos y actualizar ──
    matched_pdf_keys = set()   # para detectar instrumentos nuevos después
    deuda_gbm_matched = 0.0    # acumula deuda GBM que SÍ se mapeó a una fila
    efectivo_instr = None       # se procesa al final, después de la deuda

    for instr in instrumentos:
        fila = instr["fila"]
        nom = instr["nombre"]
        nom_n = normalizar(nom)

        old_c = valor_numerico(instr["C"])
        old_b = valor_numerico(instr["B"])

        # Flags de estado
        matched = False
        new_c = None
        new_g = None
        new_j = 0.0
        new_k = 0.0
        es_efectivo = False
        es_prestadero = False

        # ── MATCH: EFECTIVO GBM → se difiere al final del loop ──
        if "EFECTIVO" in nom_n and "GBM" in nom_n:
            efectivo_instr = instr
            continue

        # ── MATCH: DEUDA con proveedor conocido → match directo ──
        elif "DEUDA" in nom_n and instr["O"]:
            prov = normalizar(str(instr["O"]))
            fuente_key = None
            if "PRESTADERO" in prov and "prestadero" in fuentes_deuda:
                fuente_key = "prestadero"
            elif "SMART" in prov and "smart_cash" in fuentes_deuda:
                fuente_key = "smart_cash"
            elif "GBM" in prov and "gbm_deuda" in fuentes_deuda:
                fuente_key = "gbm_deuda"

            if fuente_key:
                match_fuente = fuentes_deuda[fuente_key]
                new_c = match_fuente["valor"]
                new_j = match_fuente["retiros"]
                new_k = match_fuente["depositos"]
                es_prestadero = (match_fuente["tipo"] == "prestadero")
                if es_prestadero:
                    new_g = match_fuente["interes"]
                else:
                    new_g = new_c - old_c
                matched = True
                if match_fuente["tipo"] == "gbm_deuda":
                    deuda_gbm_matched += new_c
                print(f"    🔗 DEUDA (fila {fila}) ← {fuente_key.upper()} "
                      f"(proveedor '{instr['O']}'  →  ${new_c:,.2f})")
                del fuentes_deuda[fuente_key]
            else:
                print(f"    ⬜ DEUDA (fila {fila}, prov='{instr['O']}'): "
                      f"sin fuente PDF — no se modifica")

        # ── MATCH: DEUDA sin proveedor → matching por proximidad ──
        elif "DEUDA" in nom_n and not instr["O"] and fuentes_deuda:
            match_key, match_fuente = _mejor_match_deuda(old_c, fuentes_deuda)
            if match_fuente:
                new_c = match_fuente["valor"]
                new_j = match_fuente["retiros"]
                new_k = match_fuente["depositos"]
                es_prestadero = (match_fuente["tipo"] == "prestadero")
                if es_prestadero:
                    new_g = match_fuente["interes"]
                else:
                    new_g = new_c - old_c
                matched = True
                if match_fuente["tipo"] == "gbm_deuda":
                    deuda_gbm_matched += new_c
                etiqueta = match_fuente["tipo"].upper().replace("_", " ")
                print(f"    🔗 DEUDA (fila {fila}) ← {etiqueta} "
                      f"(${new_c:,.2f}  |  anterior ${old_c:,.2f}  |  "
                      f"Δ{abs(new_c - old_c) / max(old_c, 1) * 100:.1f}%)")
                del fuentes_deuda[match_key]
            else:
                print(f"    ⚠️  DEUDA (fila {fila}, ${old_c:,.2f}): "
                      f"sin match disponible — se conserva")

        # ── MATCH: Fibras / Acciones por nombre ──
        else:
            for pdf_key, pdf_valor in pdf_port.items():
                if instrumentos_coinciden(pdf_key, nom_n):
                    new_c = pdf_valor
                    compra_neto = compras_map.get(pdf_key, 0)
                    venta_neto = ventas_map.get(pdf_key, 0)
                    new_j = venta_neto     # Retiros = dinero de ventas
                    new_k = compra_neto    # Depósitos = dinero de compras
                    # G = ganancia orgánica del mes (excluye efecto de
                    #     compras/ventas):  ΔC + retiros − depósitos
                    new_g = new_c - old_c + new_j - new_k
                    matched = True
                    matched_pdf_keys.add(pdf_key)
                    break

        # ══════════════════════════════════════════════════════
        #  Si NO hay match con PDF, NO tocar la fila en absoluto
        # ══════════════════════════════════════════════════════
        if not matched:
            print(f"    ⬜ {nom:<20s}  (sin match PDF — no se modifica)")
            continue

        # ══════════════════════════════════════════════════════
        #  Escribir SOLO los valores que corresponden
        #  ▸ B (col 2) NUNCA se toca — EXCEPTO re-compra de
        #    instrumento con B=0 (inversión original nueva)
        # ══════════════════════════════════════════════════════

        # B para fibras/acciones = B_prev + K - J (costo base se actualiza
        # con compras y ventas).  DEUDA y EFECTIVO nunca cambian B.
        if not es_efectivo and "DEUDA" not in nom_n:
            new_b = old_b + new_k - new_j
            actualizar_celda(ws, fila, 2, round(new_b, 2))
            old_b = new_b   # usar para cálculos E, F, H

        # C = Saldo Total (nuevo valor del PDF)
        actualizar_celda(ws, fila, 3, round(new_c, 2))

        # E = Ganancia Histórica = C − B
        if es_efectivo:
            actualizar_celda(ws, fila, 5, "-")
            actualizar_celda(ws, fila, 6, "-")
        else:
            new_e = new_c - old_b
            actualizar_celda(ws, fila, 5, round(new_e, 2))
            # F = Ganancia Histórica % = E / B
            new_f = (new_e / old_b) if old_b > 0 else 0.0
            actualizar_celda(ws, fila, 6, round(new_f, 10))

        # G = Ganancia / Pérdida del Mes
        if new_g is not None:
            actualizar_celda(ws, fila, 7, round(new_g, 2))

        # H = Ganancia / Pérdida del Mes %
        if es_efectivo:
            actualizar_celda(ws, fila, 8, "-")
        elif new_g is not None and old_b > 0:
            new_h = new_g / old_b
            actualizar_celda(ws, fila, 8, round(new_h, 10))

        # J = Retiros del mes,  K = Depósitos del mes
        actualizar_celda(ws, fila, 10, round(new_j, 2))
        actualizar_celda(ws, fila, 11, round(new_k, 2))

        # N = Total (= C)
        actualizar_celda(ws, fila, 14, round(new_c, 2))

        if new_g is not None:
            print(f"    ✅ {nom:<20s}  C=${new_c:>12,.2f}  G=${new_g:>10,.2f}")
        else:
            print(f"    ✅ {nom:<20s}  C=${new_c:>12,.2f}")

    # ── Procesar EFECTIVO GBM (diferido para conocer deuda_gbm_matched) ──
    if efectivo_instr and gbm:
        fila = efectivo_instr["fila"]
        old_c = valor_numerico(efectivo_instr["C"])
        gbm_total = gbm["valor_total"]
        sum_port = sum(pdf_port.values())
        new_c = round(gbm_total - sum_port - deuda_gbm_matched, 2)
        if new_c < 0:
            new_c = 0.0
        new_g = new_c - old_c
        actualizar_celda(ws, fila, 3, round(new_c, 2))
        actualizar_celda(ws, fila, 5, "-")
        actualizar_celda(ws, fila, 6, "-")
        actualizar_celda(ws, fila, 7, round(new_g, 2))
        actualizar_celda(ws, fila, 8, "-")
        actualizar_celda(ws, fila, 10, 0.0)
        actualizar_celda(ws, fila, 11, 0.0)
        actualizar_celda(ws, fila, 14, round(new_c, 2))
        print(f"    ✅ {'EFECTIVO GBM':<20s}  C=${new_c:>12,.2f}  G=${new_g:>10,.2f}")

    # ── Insertar instrumentos nuevos del portafolio GBM ──
    nuevos = []
    for pdf_key, pdf_valor in pdf_port.items():
        if pdf_key not in matched_pdf_keys:
            compra = compras_map.get(pdf_key, 0)
            venta = ventas_map.get(pdf_key, 0)
            # No insertar instrumentos sin valor ni movimientos
            if pdf_valor == 0 and compra == 0 and venta == 0:
                continue
            nuevos.append({
                "emisora": pdf_key,
                "valor_a_mercado": pdf_valor,
                "compra_neto": compra,
                "venta_neto": venta,
            })

    if nuevos:
        # Insertar ANTES de EFECTIVO GBM (para que EFECTIVO quede al final)
        if efectivo_instr:
            fila_insercion = efectivo_instr["fila"]
        else:
            fila_insercion = fila_totales
        fila_ref = instrumentos[-1]["fila"] if instrumentos else fila_header + 1
        for nuevo in nuevos:
            # insertar_instrumento inserta en fila_insercion y desplaza todo abajo
            insertar_instrumento(ws, fila_insercion, nuevo, fila_ref, periodo)
            fila_insercion += 1
            fila_totales += 1
            print(f"    ➕ NUEVO: {nuevo['emisora']:<16s}  "
                  f"C=${nuevo['valor_a_mercado']:>12,.2f}  "
                  f"K=${nuevo['compra_neto']:>10,.2f}")
        # Expandir fórmulas SUM en TOTALES para abarcar filas nuevas
        expandir_formulas_totales(ws, fila_totales)

    # ── Eliminar secciones debajo de TOTALES ──
    ultima_fila = ws.max_row
    if ultima_fila > fila_totales:
        ws.delete_rows(fila_totales + 1, ultima_fila - fila_totales)
        print(f"    🗑️  Eliminadas {ultima_fila - fila_totales} filas "
              f"debajo de TOTALES (filas {fila_totales + 1}-{ultima_fila})")

    # ── Advertir si quedan fuentes DEUDA sin mapear ──
    if fuentes_deuda:
        for key, fuente in fuentes_deuda.items():
            etiqueta = fuente["tipo"].upper().replace("_", " ")
            print(f"    ⚠️  {etiqueta} (${fuente['valor']:,.2f}) NO se pudo "
                  f"mapear a ninguna fila DEUDA. Revisar manualmente.")

    # ── NO se reescribe TOTALES, L (% Cartera), ni celdas de resumen ──
    # Esas celdas contienen fórmulas en el maestro y se actualizan solas
    # cuando el archivo se abre en Excel.

    # Periodo (forzar escritura en celdas de texto)
    if periodo:
        actualizar_celda(ws, 2, 9, f"CORTE MENSUAL {periodo['mes_nombre']}", forzar=True)
        actualizar_celda(ws, 3, 9, periodo["periodo"], forzar=True)
        # Año en K7
        k7 = ws.cell(7, 11).value
        if k7 and "\n" in str(k7):
            actualizar_celda(ws, 7, 11,
                             f"RENDIMIENTO ANUAL\n{periodo['anio']}", forzar=True)

    print(f"    ── Hoja actualizada: solo filas con match PDF modificadas.")


# ════════════════════════════════════════════════════════════
# PASO 3:  BUSCAR LA HOJA CORRECTA PARA CADA CLIENTE
# ════════════════════════════════════════════════════════════
def buscar_hoja(wb, nombre_cliente):
    """Busca la hoja del workbook que coincida con el nombre del cliente."""
    nc = nombre_cliente.upper().strip()
    # Match exacto
    for s in wb.sheetnames:
        if s.upper().strip() == nc:
            return s
    # Match parcial
    for s in wb.sheetnames:
        if nc in s.upper() or s.upper().strip() in nc:
            return s
    # Match por apellidos
    partes = nc.split()
    for s in wb.sheetnames:
        su = s.upper()
        coincidencias = sum(1 for p in partes if p in su)
        if coincidencias >= 2:
            return s
    return None


# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════
def main():
    print("╔══════════════════════════════════════════════════╗")
    print("║  CONSOLIDADOR DE ESTADOS DE CUENTA              ║")
    print("╚══════════════════════════════════════════════════╝\n")

    # ── 1. Buscar archivo maestro ──
    maestros = [f for f in os.listdir(CARPETA_MAESTRO)
                if f.endswith(".xlsx") and not f.startswith("~") and not f.startswith(".")]
    if not maestros:
        print("❌ No se encontró ningún archivo .xlsx en 00_Maestro_Anterior/")
        print("   Coloca ahí el Excel del mes anterior y vuelve a ejecutar.")
        return
    # Seleccionar el más reciente por fecha de modificación
    maestros.sort(key=lambda f: os.path.getmtime(os.path.join(CARPETA_MAESTRO, f)), reverse=True)
    maestro_archivo = maestros[0]
    maestro_ruta = os.path.join(CARPETA_MAESTRO, maestro_archivo)
    print(f"📂 Maestro: {maestro_archivo}")
    if len(maestros) > 1:
        print(f"   (Se encontraron {len(maestros)} archivos, usando el más reciente)")
    print()

    # ── 2. Extraer datos de PDFs ──
    print("─── Extrayendo datos de PDFs ───")
    clientes = extraer_todos_los_pdfs()
    if not clientes:
        print("\n⚠️  No se encontraron PDFs en PDFs_Origen/")
        return
    print(f"\n   Clientes detectados: {len(clientes)}")
    for n in sorted(clientes):
        tipos = []
        if clientes[n]["gbm"]:        tipos.append("GBM")
        if clientes[n]["smart_cash"]: tipos.append("Smart Cash")
        if clientes[n]["prestadero"]: tipos.append("Prestadero")
        print(f"     • {n}  ({', '.join(tipos)})")

    # ── 3. Copiar maestro a salida ──
    # Determinar nombre del mes para el archivo de salida
    any_periodo = None
    for c in clientes.values():
        if c["periodo"]:
            any_periodo = c["periodo"]
            break
    mes_nombre = any_periodo["mes_nombre"] if any_periodo else "ACTUALIZADO"
    anio = any_periodo["anio"] if any_periodo else ""
    salida_nombre = f"ESTADOS DE CUENTA {mes_nombre} {anio}.xlsx".strip()
    salida_ruta = os.path.join(CARPETA_SALIDA, salida_nombre)

    shutil.copy2(maestro_ruta, salida_ruta)
    print(f"\n📋 Copiando maestro → {salida_nombre}")

    # ── 4. Abrir copia y actualizar hojas ──
    wb = load_workbook(salida_ruta)
    print(f"   Hojas en maestro: {wb.sheetnames}\n")

    print("─── Consolidando datos ───")
    clientes_actualizados = 0
    clientes_sin_hoja = []

    for nombre_cliente, datos in sorted(clientes.items()):
        hoja = buscar_hoja(wb, nombre_cliente)
        if hoja:
            print(f"\n  🔄 {nombre_cliente}  →  Hoja: '{hoja}'")
            ws = wb[hoja]
            try:
                actualizar_hoja(ws, datos, hoja)
                clientes_actualizados += 1
            except Exception as e:
                print(f"    ❌ Error actualizando: {e}")
                import traceback; traceback.print_exc()
        else:
            clientes_sin_hoja.append(nombre_cliente)
            print(f"\n  ⚠️  {nombre_cliente} — No tiene hoja en el maestro (se omite)")

    # ── 5. Guardar ──
    wb.save(salida_ruta)

    # ── Resumen final ──
    print(f"\n{'═'*52}")
    print(f"  ✅ Archivo generado: {salida_nombre}")
    print(f"     Ubicación: {salida_ruta}")
    print(f"     Clientes actualizados: {clientes_actualizados}")
    if clientes_sin_hoja:
        print(f"     ⚠️  Sin hoja en maestro: {', '.join(clientes_sin_hoja)}")
    print(f"{'═'*52}")


if __name__ == "__main__":
    main()
