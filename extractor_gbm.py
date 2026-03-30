import pdfplumber
import pandas as pd
import os
import re
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# RUTAS (relativas al directorio del script)
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
carpeta_pdfs = os.path.join(BASE_DIR, "PDFs_Origen")
carpeta_salida = os.path.join(BASE_DIR, "Resultados_Excel")
archivo_salida = os.path.join(carpeta_salida, "Reporte_Consolidado.xlsx")

os.makedirs(carpeta_salida, exist_ok=True)

print("Iniciando la lectura inteligente de PDFs...\n")


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================
def extraer_numeros(texto):
    """Extrae números con punto decimal como '1,022.62'."""
    numeros = re.findall(r'[\d,]+\.\d+', texto)
    return [float(n.replace(',', '')) for n in numeros]


def extraer_todos_numeros(texto):
    """Extrae TODOS los números (enteros y decimales)."""
    numeros = re.findall(r'[\d,]+\.?\d*', texto)
    return [float(n.replace(',', '')) for n in numeros if n]


def extraer_numero_despues_de(texto, clave):
    """Extrae el primer número decimal que aparece DESPUÉS de una clave."""
    idx = texto.find(clave)
    if idx == -1:
        return None
    subtexto = texto[idx + len(clave):]
    nums = extraer_numeros(subtexto)
    return nums[0] if nums else None


def extraer_nombre_cliente(pdf, plataforma):
    """Extrae el nombre del cliente desde la primera página del PDF."""
    texto = pdf.pages[0].extract_text()
    if not texto:
        return "Desconocido"

    lineas = texto.split('\n')

    if plataforma == "Prestadero":
        for linea in lineas:
            if "Periodo:" in linea and "Estado de Cuenta" not in linea:
                nombre = re.split(r'\s+Periodo:', linea)[0].strip()
                return nombre.upper()
    else:
        for linea in lineas:
            if "Contrato:" in linea:
                parte = re.split(r'\s+Contrato:', linea)[0].strip()
                if "PUBLICO EN GENERAL - " in parte:
                    parte = parte.replace("PUBLICO EN GENERAL - ", "")
                return parte.upper()

    return "Desconocido"


def detectar_plataforma(texto):
    """Detecta si el PDF es Prestadero o GBM."""
    if "Prestadero" in texto or "PRESTADERO" in texto:
        return "Prestadero"
    return "GBM"


def es_smart_cash(texto):
    """Detecta si un PDF de GBM es Smart Cash (sin renta variable)."""
    for linea in texto.split('\n'):
        if "RENTA VARIABLE" in linea and "VALORES EN CORTO" not in linea:
            nums = extraer_numeros(linea)
            if len(nums) >= 2 and nums[1] > 0:
                return False
    return True


def extraer_saldo_anterior(lineas_p1):
    """Extrae el saldo anterior del RESUMEN DEL PORTAFOLIO (primer número de VALOR DEL PORTAFOLIO)."""
    for linea in lineas_p1:
        if "VALOR DEL PORTAFOLIO" in linea and "TOTAL" not in linea:
            nums = extraer_numeros(linea)
            if len(nums) >= 1:
                return nums[0]
    return 0.0


def extraer_portafolio_gbm(pdf):
    """Extrae la tabla ACCIONES del DESGLOSE DEL PORTAFOLIO.
    Columnas: Emisora, Tít Ant, Tít Act, Costo Total, Precio Merc Ant, Precio Merc Act, Valor a Mercado.
    Formato línea: 'FIBRAPL 14 5 10 0 72.500000 725.00 78.070000 74.190000 780.70 55.70 2.88'
    Índices nums:   [0]tit_ant [1]tit_act [2]tit_prest [3]costo_prom [4]costo_total
                    [5]precio_ant [6]precio_act [7]valor_merc [8]plusvalia [9]%cart
    """
    portafolio = []
    en_desglose = False
    en_acciones = False

    for pagina in pdf.pages:
        texto = pagina.extract_text()
        if not texto:
            continue
        for linea in texto.split('\n'):
            linea_strip = linea.strip()
            linea_upper = linea_strip.upper()

            if "DESGLOSE DEL PORTAFOLIO" in linea_upper:
                en_desglose = True
                continue
            if not en_desglose:
                continue
            if linea_upper == "ACCIONES":
                en_acciones = True
                continue
            if en_acciones and ("EMISORA" in linea_upper or "MES ANTERIOR" in linea_upper
                                or "MES ACTUAL" in linea_upper or "EN PR" in linea_upper):
                continue
            if en_acciones and linea_upper.startswith("TOTAL"):
                en_acciones = False
                continue
            if en_desglose and linea_upper in (
                "DESGLOSE DE MOVIMIENTOS", "RENDIMIENTO DEL PORTAFOLIO", "EFECTIVO",
            ):
                en_desglose = False
                en_acciones = False
                continue
            if not en_acciones:
                continue

            match = re.match(r'^([A-Z]+(?:\s+\d+)?)\s+', linea_strip)
            if match:
                try:
                    emisora = match.group(1).strip()
                    parte_nums = linea_strip[match.end():]
                    nums = extraer_todos_numeros(parte_nums)
                    if len(nums) >= 8:
                        portafolio.append({
                            "Emisora": emisora,
                            "Títulos Mes Anterior": int(nums[0]),
                            "Títulos Mes Actual": int(nums[1]),
                            "Costo Total": nums[4],
                            "Precio Mercado Mes Anterior": nums[5],
                            "Precio Mercado Mes Actual": nums[6],
                            "Valor a Mercado": nums[7],
                        })
                except (ValueError, IndexError):
                    continue
    return portafolio


def extraer_deuda_gbm(pdf):
    """Extrae la tabla DEUDA EN REPORTO del DESGLOSE DEL PORTAFOLIO.
    Formato: 'BI 260528 0 9,872 4.50 4 4 9.768849 9.706696 96,438.08 0.00 99.99'
    Índices: [0]tit_ant [1]tit_act [2]tasa [3]plazo [4]dias_vencer
             [5]precio_ant [6]precio_act [7]valor_reporto [8]premio_dev [9]%cart
    """
    deuda = []
    en_desglose = False
    en_deuda = False

    for pagina in pdf.pages:
        texto = pagina.extract_text()
        if not texto:
            continue
        for linea in texto.split('\n'):
            linea_strip = linea.strip()
            linea_upper = linea_strip.upper()

            if "DESGLOSE DEL PORTAFOLIO" in linea_upper:
                en_desglose = True
                continue
            if not en_desglose:
                continue
            if "DEUDA EN REPORTO" in linea_upper and "TOTAL" not in linea_upper:
                en_deuda = True
                continue
            if en_deuda and ("EMISORA" in linea_upper or "ANTERIOR" in linea_upper):
                continue
            if en_deuda and linea_upper.startswith("TOTAL"):
                en_deuda = False
                continue
            if en_desglose and linea_upper in (
                "RENTA VARIABLE", "EFECTIVO", "DESGLOSE DE MOVIMIENTOS",
                "RENDIMIENTO DEL PORTAFOLIO",
            ):
                en_desglose = False
                en_deuda = False
                continue
            if not en_deuda:
                continue

            match = re.match(r'^([A-Z]+\s+\d+)\s+', linea_strip)
            if match:
                try:
                    emisora = match.group(1).strip()
                    parte_nums = linea_strip[match.end():]
                    nums = extraer_todos_numeros(parte_nums)
                    if len(nums) >= 8:
                        deuda.append({
                            "Emisora": emisora,
                            "Títulos Mes Anterior": int(nums[0]),
                            "Títulos Mes Actual": int(nums[1]),
                            "Tasa": nums[2],
                            "Valor del Reporto": nums[7],
                            "% Cartera": nums[9] if len(nums) >= 10 else 0.0,
                        })
                except (ValueError, IndexError):
                    continue
    return deuda


def extraer_movimientos_acciones(pdf):
    """Extrae compras y ventas de acciones del DESGLOSE DE MOVIMIENTOS."""
    movimientos = []
    en_movimientos = False

    for pagina in pdf.pages:
        texto = pagina.extract_text()
        if not texto:
            continue
        for linea in texto.split('\n'):
            linea_upper = linea.strip().upper()

            if "DESGLOSE DE MOVIMIENTOS" in linea_upper:
                en_movimientos = True
                continue
            if en_movimientos and linea_upper in (
                "RENDIMIENTO DEL PORTAFOLIO", "COMPOSICIÓN FISCAL INFORMATIVA",
            ):
                en_movimientos = False
                continue
            if not en_movimientos:
                continue
            if "Compra de Acciones" not in linea and "Venta de Acciones" not in linea:
                continue

            try:
                linea_strip = linea.strip()
                if "Compra de Acciones" in linea:
                    operacion = "Compra"
                    idx = linea.find("Compra de Acciones.")
                    resto = linea[idx + len("Compra de Acciones."):].strip()
                else:
                    operacion = "Venta"
                    idx = linea.find("Venta de Acciones.")
                    resto = linea[idx + len("Venta de Acciones."):].strip()

                fecha_match = re.match(r'(\d{2}/\d{2})', linea_strip)
                fecha = fecha_match.group(1) if fecha_match else ""

                em_match = re.match(r'^([A-Z]+(?:\s+\d+)?)\s+', resto)
                if em_match:
                    emisora = em_match.group(1).strip()
                    parte_nums = resto[em_match.end():]
                    nums = extraer_todos_numeros(parte_nums)
                    # [títulos, precio, tasa, rend, plazo, comisión, interés, impuesto, neto, saldo]
                    titulos = int(nums[0]) if len(nums) >= 1 else 0
                    precio = nums[1] if len(nums) >= 2 else 0.0
                    comision = nums[2] if len(nums) >= 3 else 0.0
                    neto = nums[5] if len(nums) >= 6 else 0.0

                    movimientos.append({
                        "Fecha": fecha,
                        "Operación": operacion,
                        "Emisora": emisora,
                        "Títulos": titulos,
                        "Precio Unitario": precio,
                        "Comisión": comision,
                        "Neto": neto,
                    })
            except Exception:
                continue
    return movimientos


def extraer_movimientos_efectivo_smart_cash(pdf):
    """Extrae depósitos y retiros del DESGLOSE DE MOVIMIENTOS de Smart Cash."""
    movimientos = []
    en_movimientos = False

    for pagina in pdf.pages:
        texto = pagina.extract_text()
        if not texto:
            continue
        for linea in texto.split('\n'):
            linea_upper = linea.strip().upper()

            if "DESGLOSE DE MOVIMIENTOS" in linea_upper:
                en_movimientos = True
                continue
            if en_movimientos and linea_upper in (
                "RENDIMIENTO DEL PORTAFOLIO", "COMPOSICIÓN FISCAL INFORMATIVA",
            ):
                en_movimientos = False
                continue
            if not en_movimientos:
                continue
            if "DEPOSITO" not in linea_upper and "RETIRO" not in linea_upper:
                continue

            try:
                linea_strip = linea.strip()
                fecha_match = re.match(r'(\d{2}/\d{2})', linea_strip)
                fecha = fecha_match.group(1) if fecha_match else ""

                if "DEPOSITO" in linea_upper:
                    operacion = "Depósito"
                else:
                    operacion = "Retiro"

                # El neto es el penúltimo número, el saldo es el último
                nums = extraer_numeros(linea_strip)
                # Formato: ... comisión interés impuesto NETO saldo
                # Los últimos 2 son neto y saldo
                monto = nums[-2] if len(nums) >= 2 else (nums[0] if nums else 0.0)

                movimientos.append({
                    "Fecha": fecha,
                    "Operación": operacion,
                    "Monto": monto,
                })
            except Exception:
                continue
    return movimientos


# ─── Formateo de Excel ───
TITULO_FONT = Font(bold=True, size=12, color="FFFFFF")
TITULO_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
)


def escribir_seccion(writer, hoja, fila, titulo, df):
    """Escribe un título con formato y un DataFrame debajo. Retorna la fila siguiente."""
    ws = writer.sheets[hoja]

    # Título con fondo azul
    celda = ws.cell(row=fila + 1, column=1, value=titulo)
    celda.font = TITULO_FONT
    celda.fill = TITULO_FILL
    celda.alignment = Alignment(horizontal="left")
    # Extender el fondo del título a todas las columnas del df
    for c in range(2, len(df.columns) + 1):
        ws.cell(row=fila + 1, column=c).fill = TITULO_FILL

    df.to_excel(writer, sheet_name=hoja, startrow=fila + 1, index=False)
    return fila + len(df) + 3


def escribir_fila_total(writer, hoja, fila, label, valor):
    """Escribe una fila de total con formato destacado."""
    ws = writer.sheets[hoja]
    c1 = ws.cell(row=fila + 1, column=1, value=label)
    c1.font = TOTAL_FONT
    c1.fill = TOTAL_FILL
    c2 = ws.cell(row=fila + 1, column=2, value=valor)
    c2.font = TOTAL_FONT
    c2.fill = TOTAL_FILL
    c2.number_format = '#,##0.00'
    return fila + 2


def ajustar_columnas(ws):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)


# ============================================================
# PASO 1: Procesar cada PDF y agrupar por cliente
# ============================================================
clientes = defaultdict(lambda: {"gbm": None, "smart_cash": None, "prestadero": None})

for archivo in sorted(os.listdir(carpeta_pdfs)):
    if not archivo.lower().endswith(".pdf"):
        continue

    ruta_completa = os.path.join(carpeta_pdfs, archivo)
    print(f"  Procesando: {archivo}")

    try:
        with pdfplumber.open(ruta_completa) as pdf:
            texto_pagina1 = pdf.pages[0].extract_text() or ""
            texto_completo = texto_pagina1
            for pagina in pdf.pages[1:]:
                t = pagina.extract_text()
                if t:
                    texto_completo += "\n" + t

            plataforma = detectar_plataforma(texto_completo)
            nombre_cliente = extraer_nombre_cliente(pdf, plataforma)
            lineas_p1 = texto_pagina1.split('\n')

            print(f"    Cliente: {nombre_cliente} | Plataforma: {plataforma}")

            # =============================================
            # PRESTADERO
            # =============================================
            if plataforma == "Prestadero":
                abonos = 0.0
                retiros = 0.0
                interes = 0.0
                valor_cuenta = 0.0

                for linea in lineas_p1:
                    try:
                        if "Abonos:" in linea and "Cuenta Abonos:" not in linea:
                            val = extraer_numero_despues_de(linea, "Abonos:")
                            if val is not None:
                                abonos = val
                        if "Valor de la Cuenta:" in linea:
                            val = extraer_numero_despues_de(linea, "Valor de la Cuenta:")
                            if val is not None:
                                valor_cuenta = val
                        if "Interés Recibido" in linea or "Interes Recibido" in linea:
                            nums = extraer_numeros(linea)
                            if nums:
                                interes = nums[0]
                        if "Retiros:" in linea and "Detalle" not in linea:
                            val = extraer_numero_despues_de(linea, "Retiros:")
                            if val is not None:
                                retiros = val
                    except Exception:
                        continue

                clientes[nombre_cliente]["prestadero"] = {
                    "resumen": pd.DataFrame([{
                        "Plataforma": "Prestadero",
                        "Abonos": abonos,
                        "Retiros": retiros,
                        "Interés Recibido": interes,
                        "Valor de la Cuenta": valor_cuenta,
                    }]),
                }

            # =============================================
            # GBM (Regular o Smart Cash)
            # =============================================
            else:
                entradas = 0.0
                salidas = 0.0
                valor_total = 0.0
                saldo_anterior = 0.0

                saldo_anterior = extraer_saldo_anterior(lineas_p1)

                for linea in lineas_p1:
                    try:
                        if "ENTRADAS DE EFECTIVO" in linea:
                            nums = extraer_numeros(linea)
                            if nums:
                                entradas = nums[-1]
                        elif "SALIDAS DE EFECTIVO" in linea:
                            nums = extraer_numeros(linea)
                            if nums:
                                salidas = nums[-1]
                        elif "VALOR DEL PORTAFOLIO" in linea and "TOTAL" not in linea:
                            nums = extraer_numeros(linea)
                            if len(nums) >= 2:
                                valor_total = nums[1]
                            elif len(nums) == 1:
                                valor_total = nums[0]
                    except Exception:
                        continue

                smart_cash = es_smart_cash(texto_pagina1)
                portafolio = []
                deuda = []
                movimientos = []
                movimientos_efectivo = []

                try:
                    deuda = extraer_deuda_gbm(pdf)
                except Exception as e:
                    print(f"    ⚠️ Error extrayendo deuda: {e}")

                if smart_cash:
                    try:
                        movimientos_efectivo = extraer_movimientos_efectivo_smart_cash(pdf)
                    except Exception as e:
                        print(f"    ⚠️ Error extrayendo movimientos Smart Cash: {e}")
                else:
                    try:
                        portafolio = extraer_portafolio_gbm(pdf)
                    except Exception as e:
                        print(f"    ⚠️ Error extrayendo portafolio: {e}")
                    try:
                        movimientos = extraer_movimientos_acciones(pdf)
                    except Exception as e:
                        print(f"    ⚠️ Error extrayendo movimientos: {e}")

                tipo = "smart_cash" if smart_cash else "gbm"
                print(f"    Tipo: {'Smart Cash' if smart_cash else 'GBM Regular'}")

                datos = {
                    "resumen": pd.DataFrame([{
                        "Plataforma": "GBM Smart Cash" if smart_cash else "GBM",
                        "Saldo Anterior": saldo_anterior,
                        "Entradas de Efectivo": entradas,
                        "Salidas de Efectivo": salidas,
                        "Valor Total del Portafolio": valor_total,
                    }]),
                    "deuda": deuda,
                    "portafolio": portafolio,
                    "movimientos": movimientos,
                    "movimientos_efectivo": movimientos_efectivo,
                }

                clientes[nombre_cliente][tipo] = datos

    except Exception as e:
        print(f"    ❌ Error procesando {archivo}: {e}")


# ============================================================
# PASO 2: Generar un solo Excel con una hoja por cliente
# ============================================================
if not clientes:
    print("⚠️ No se encontraron PDFs.")
else:
    with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
        for nombre_cliente, cuentas in sorted(clientes.items()):
            hoja = nombre_cliente[:31]
            print(f"\n  Escribiendo hoja: {hoja}")

            pd.DataFrame().to_excel(writer, sheet_name=hoja, index=False)
            fila = 0

            # ─── GBM Regular ───
            if cuentas["gbm"]:
                gbm = cuentas["gbm"]
                fila = escribir_seccion(writer, hoja, fila, "RESUMEN GBM", gbm["resumen"])

                if gbm["portafolio"]:
                    df = pd.DataFrame(gbm["portafolio"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "PORTAFOLIO (ACCIONES / FIBRAS)", df)

                if gbm["deuda"]:
                    df = pd.DataFrame(gbm["deuda"])
                    fila = escribir_seccion(writer, hoja, fila, "DEUDA", df)

                if gbm["movimientos"]:
                    df_mov = pd.DataFrame(gbm["movimientos"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "COMPRA / VENTA DE ACCIONES", df_mov)

                    # Totales de compras y ventas
                    compras = df_mov[df_mov["Operación"] == "Compra"]
                    ventas = df_mov[df_mov["Operación"] == "Venta"]
                    total_compras = compras["Neto"].sum() if len(compras) > 0 else 0.0
                    total_ventas = ventas["Neto"].sum() if len(ventas) > 0 else 0.0
                    fila = escribir_fila_total(writer, hoja, fila,
                                               "COMPRA TOTAL DE ACCIONES", total_compras)
                    fila = escribir_fila_total(writer, hoja, fila,
                                               "VENTA TOTAL DE ACCIONES", total_ventas)
                    fila += 1  # Espacio extra después de totales

            # ─── GBM Smart Cash ───
            if cuentas["smart_cash"]:
                sc = cuentas["smart_cash"]
                fila = escribir_seccion(writer, hoja, fila, "RESUMEN SMART CASH", sc["resumen"])

                if sc["deuda"]:
                    df = pd.DataFrame(sc["deuda"])
                    fila = escribir_seccion(writer, hoja, fila, "DEUDA SMART CASH", df)

                if sc.get("movimientos_efectivo"):
                    df_efec = pd.DataFrame(sc["movimientos_efectivo"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "MOVIMIENTOS DE EFECTIVO SMART CASH", df_efec)

                    # Totales depósitos y retiros
                    depositos = df_efec[df_efec["Operación"] == "Depósito"]
                    retiros = df_efec[df_efec["Operación"] == "Retiro"]
                    total_dep = depositos["Monto"].sum() if len(depositos) > 0 else 0.0
                    total_ret = retiros["Monto"].sum() if len(retiros) > 0 else 0.0
                    fila = escribir_fila_total(writer, hoja, fila,
                                               "TOTAL DEPÓSITOS", total_dep)
                    fila = escribir_fila_total(writer, hoja, fila,
                                               "TOTAL RETIROS", total_ret)
                    fila += 1

            # ─── Prestadero ───
            if cuentas["prestadero"]:
                prest = cuentas["prestadero"]
                fila = escribir_seccion(writer, hoja, fila, "RESUMEN PRESTADERO",
                                        prest["resumen"])

            # Ajustar columnas
            ajustar_columnas(writer.sheets[hoja])

    print(f"\n{'='*50}")
    print(f"✅ Archivo generado: {archivo_salida}")
    print(f"   Hojas creadas: {len(clientes)}")
    for nombre in sorted(clientes.keys()):
        tipos = []
        if clientes[nombre]["gbm"]:
            tipos.append("GBM")
        if clientes[nombre]["smart_cash"]:
            tipos.append("Smart Cash")
        if clientes[nombre]["prestadero"]:
            tipos.append("Prestadero")
        print(f"     - {nombre} ({', '.join(tipos)})")
