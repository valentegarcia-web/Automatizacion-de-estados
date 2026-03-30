#!/usr/bin/env python3
"""
app.py — Streamlit App
Automatizacion de Estados de Cuenta (GBM / Prestadero)
"""

import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from collections import defaultdict
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ════════════════════════════════════════════════════════════
# CONFIGURACION DE PAGINA
# ════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Estados de Cuenta - GBM",
    page_icon="📊",
    layout="wide",
)

# ════════════════════════════════════════════════════════════
# CSS PERSONALIZADO
# ════════════════════════════════════════════════════════════
st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1rem 0;
    }
    .step-header {
        background: linear-gradient(90deg, #2F5496, #4472C4);
        color: white;
        padding: 0.8rem 1.2rem;
        border-radius: 8px;
        margin-bottom: 1rem;
        font-size: 1.1rem;
    }
    .info-box {
        background-color: #f0f4ff;
        border-left: 4px solid #2F5496;
        padding: 0.8rem 1rem;
        border-radius: 0 8px 8px 0;
        margin: 0.5rem 0;
    }
    .success-box {
        background-color: #f0fff4;
        border-left: 4px solid #28a745;
        padding: 0.8rem 1rem;
        border-radius: 0 8px 8px 0;
    }
    .download-section {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        border: 1px solid #dee2e6;
        margin-top: 1rem;
    }
    div[data-testid="stFileUploader"] {
        border: 2px dashed #4472C4;
        border-radius: 10px;
        padding: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# FUNCIONES DE EXTRACCION
# ════════════════════════════════════════════════════════════
def extraer_numeros(texto):
    nums = re.findall(r'[\d,]+\.\d+', texto)
    return [float(n.replace(',', '')) for n in nums]


def extraer_todos_numeros(texto):
    nums = re.findall(r'[\d,]+\.?\d*', texto)
    return [float(n.replace(',', '')) for n in nums if n]


def extraer_numero_despues_de(texto, clave):
    idx = texto.find(clave)
    if idx == -1:
        return None
    sub = texto[idx + len(clave):]
    nums = extraer_numeros(sub)
    return nums[0] if nums else None


def extraer_nombre_cliente(pdf, plataforma):
    texto = pdf.pages[0].extract_text() or ""
    lineas = texto.split('\n')
    if plataforma == "Prestadero":
        for linea in lineas:
            if "Periodo:" in linea and "Estado de Cuenta" not in linea:
                return re.split(r'\s+Periodo:', linea)[0].strip().upper()
    else:
        for linea in lineas:
            if "Contrato:" in linea:
                parte = re.split(r'\s+Contrato:', linea)[0].strip()
                parte = parte.replace("PUBLICO EN GENERAL - ", "")
                return parte.upper()
    return "DESCONOCIDO"


def detectar_plataforma(texto):
    return "Prestadero" if ("Prestadero" in texto or "PRESTADERO" in texto) else "GBM"


def es_smart_cash(texto):
    for linea in texto.split('\n'):
        if "RENTA VARIABLE" in linea and "VALORES EN CORTO" not in linea:
            nums = extraer_numeros(linea)
            if len(nums) >= 2 and nums[1] > 0:
                return False
    return True


def extraer_saldo_anterior(lineas):
    for l in lineas:
        if "VALOR DEL PORTAFOLIO" in l and "TOTAL" not in l:
            nums = extraer_numeros(l)
            if nums:
                return nums[0]
    return 0.0


def extraer_portafolio_gbm(pdf):
    portafolio = []
    en_desglose = en_acciones = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu:
                en_desglose = True; continue
            if not en_desglose: continue
            if lu == "ACCIONES":
                en_acciones = True; continue
            if en_acciones and ("EMISORA" in lu or "MES ANTERIOR" in lu or "EN PR" in lu):
                continue
            if en_acciones and lu.startswith("TOTAL"):
                en_acciones = False; continue
            if lu in ("DESGLOSE DE MOVIMIENTOS", "RENDIMIENTO DEL PORTAFOLIO", "EFECTIVO"):
                en_desglose = en_acciones = False; continue
            if not en_acciones: continue
            m = re.match(r'^([A-Z]+(?:\s+\d+)?)\s+', ls)
            if m:
                try:
                    emisora = m.group(1).strip()
                    nums = extraer_todos_numeros(ls[m.end():])
                    if len(nums) >= 8:
                        portafolio.append({
                            "Emisora": emisora,
                            "Titulos Mes Anterior": int(nums[0]),
                            "Titulos Mes Actual": int(nums[1]),
                            "Costo Total": nums[4],
                            "Precio Mercado Mes Anterior": nums[5],
                            "Precio Mercado Mes Actual": nums[6],
                            "Valor a Mercado": nums[7],
                        })
                except (ValueError, IndexError):
                    continue
    return portafolio


def extraer_deuda_gbm(pdf):
    deuda = []
    en_desglose = en_deuda = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu:
                en_desglose = True; continue
            if not en_desglose: continue
            if "DEUDA EN REPORTO" in lu and "TOTAL" not in lu:
                en_deuda = True; continue
            if en_deuda and ("EMISORA" in lu or "ANTERIOR" in lu):
                continue
            if en_deuda and lu.startswith("TOTAL"):
                en_deuda = False; continue
            if lu in ("RENTA VARIABLE", "EFECTIVO", "DESGLOSE DE MOVIMIENTOS",
                       "RENDIMIENTO DEL PORTAFOLIO"):
                en_desglose = en_deuda = False; continue
            if not en_deuda: continue
            m = re.match(r'^([A-Z]+\s+\d+)\s+', ls)
            if m:
                try:
                    emisora = m.group(1).strip()
                    nums = extraer_todos_numeros(ls[m.end():])
                    if len(nums) >= 8:
                        deuda.append({
                            "Emisora": emisora,
                            "Titulos Mes Anterior": int(nums[0]),
                            "Titulos Mes Actual": int(nums[1]),
                            "Tasa": nums[2],
                            "Valor del Reporto": nums[7],
                            "% Cartera": nums[9] if len(nums) >= 10 else 0.0,
                        })
                except (ValueError, IndexError):
                    continue
    return deuda


def extraer_movimientos_acciones(pdf):
    movimientos = []
    en_mov = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            lu = l.strip().upper()
            if "DESGLOSE DE MOVIMIENTOS" in lu:
                en_mov = True; continue
            if en_mov and lu in ("RENDIMIENTO DEL PORTAFOLIO",
                                  "COMPOSICIÓN FISCAL INFORMATIVA"):
                en_mov = False; continue
            if not en_mov: continue
            if "Compra de Acciones" not in l and "Venta de Acciones" not in l:
                continue
            try:
                ls = l.strip()
                if "Compra de Acciones" in l:
                    op = "Compra"
                    resto = l[l.find("Compra de Acciones.") +
                              len("Compra de Acciones."):].strip()
                else:
                    op = "Venta"
                    resto = l[l.find("Venta de Acciones.") +
                              len("Venta de Acciones."):].strip()
                fm = re.match(r'(\d{2}/\d{2})', ls)
                fecha = fm.group(1) if fm else ""
                em = re.match(r'^([A-Z]+(?:\s+\d+)?)\s+', resto)
                if em:
                    emisora = em.group(1).strip()
                    nums = extraer_todos_numeros(resto[em.end():])
                    movimientos.append({
                        "Fecha": fecha, "Operacion": op, "Emisora": emisora,
                        "Titulos": int(nums[0]) if nums else 0,
                        "Precio Unitario": nums[1] if len(nums) >= 2 else 0,
                        "Comision": nums[2] if len(nums) >= 3 else 0,
                        "Neto": nums[5] if len(nums) >= 6 else 0,
                    })
            except Exception:
                continue
    return movimientos


def extraer_movimientos_efectivo_smart_cash(pdf):
    movimientos = []
    en_mov = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            lu = l.strip().upper()
            if "DESGLOSE DE MOVIMIENTOS" in lu:
                en_mov = True; continue
            if en_mov and lu in ("RENDIMIENTO DEL PORTAFOLIO",
                                  "COMPOSICIÓN FISCAL INFORMATIVA"):
                en_mov = False; continue
            if not en_mov: continue
            if "DEPOSITO" not in lu and "RETIRO" not in lu:
                continue
            try:
                ls = l.strip()
                fecha_match = re.match(r'(\d{2}/\d{2})', ls)
                fecha = fecha_match.group(1) if fecha_match else ""
                operacion = "Deposito" if "DEPOSITO" in lu else "Retiro"
                nums = extraer_numeros(ls)
                monto = nums[-2] if len(nums) >= 2 else (nums[0] if nums else 0.0)
                movimientos.append({
                    "Fecha": fecha, "Operacion": operacion, "Monto": monto,
                })
            except Exception:
                continue
    return movimientos


# ════════════════════════════════════════════════════════════
# FORMATEO EXCEL
# ════════════════════════════════════════════════════════════
TITULO_FONT = Font(bold=True, size=12, color="FFFFFF")
TITULO_FILL = PatternFill(start_color="2F5496", end_color="2F5496",
                           fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                          fill_type="solid")


def escribir_seccion(writer, hoja, fila, titulo, df):
    ws = writer.sheets[hoja]
    celda = ws.cell(row=fila + 1, column=1, value=titulo)
    celda.font = TITULO_FONT
    celda.fill = TITULO_FILL
    celda.alignment = Alignment(horizontal="left")
    for c in range(2, len(df.columns) + 1):
        ws.cell(row=fila + 1, column=c).fill = TITULO_FILL
    df.to_excel(writer, sheet_name=hoja, startrow=fila + 1, index=False)
    return fila + len(df) + 3


def escribir_fila_total(writer, hoja, fila, label, valor):
    ws = writer.sheets[hoja]
    c1 = ws.cell(row=fila + 1, column=1, value=label)
    c1.font = TOTAL_FONT; c1.fill = TOTAL_FILL
    c2 = ws.cell(row=fila + 1, column=2, value=valor)
    c2.font = TOTAL_FONT; c2.fill = TOTAL_FILL
    c2.number_format = '#,##0.00'
    return fila + 2


def ajustar_columnas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)


# ════════════════════════════════════════════════════════════
# PROCESAR PDFs
# ════════════════════════════════════════════════════════════
def procesar_pdfs(archivos_pdf):
    clientes = defaultdict(lambda: {
        "gbm": None, "smart_cash": None, "prestadero": None,
    })
    log = []

    for archivo in archivos_pdf:
        nombre_archivo = archivo.name
        log.append(f"Procesando: {nombre_archivo}")
        try:
            with pdfplumber.open(archivo) as pdf:
                texto_p1 = pdf.pages[0].extract_text() or ""
                texto_completo = texto_p1
                for p in pdf.pages[1:]:
                    t = p.extract_text()
                    if t:
                        texto_completo += "\n" + t

                plataforma = detectar_plataforma(texto_completo)
                nombre = extraer_nombre_cliente(pdf, plataforma)
                lineas_p1 = texto_p1.split('\n')
                log.append(f"  -> Cliente: {nombre} | {plataforma}")

                if plataforma == "Prestadero":
                    abonos = retiros = interes = valor = 0.0
                    for l in lineas_p1:
                        try:
                            if "Abonos:" in l and "Cuenta Abonos:" not in l:
                                v = extraer_numero_despues_de(l, "Abonos:")
                                if v is not None: abonos = v
                            if "Valor de la Cuenta:" in l:
                                v = extraer_numero_despues_de(l, "Valor de la Cuenta:")
                                if v is not None: valor = v
                            if "Interés Recibido" in l or "Interes Recibido" in l:
                                ns = extraer_numeros(l)
                                if ns: interes = ns[0]
                            if "Retiros:" in l and "Detalle" not in l:
                                v = extraer_numero_despues_de(l, "Retiros:")
                                if v is not None: retiros = v
                        except Exception:
                            continue
                    clientes[nombre]["prestadero"] = {
                        "resumen": pd.DataFrame([{
                            "Plataforma": "Prestadero", "Abonos": abonos,
                            "Retiros": retiros, "Interes Recibido": interes,
                            "Valor de la Cuenta": valor,
                        }]),
                        "abonos": abonos, "retiros": retiros,
                        "interes": interes, "valor": valor,
                    }
                else:
                    entradas = salidas = valor_total = saldo_ant = 0.0
                    saldo_ant = extraer_saldo_anterior(lineas_p1)
                    for l in lineas_p1:
                        try:
                            if "ENTRADAS DE EFECTIVO" in l:
                                ns = extraer_numeros(l)
                                if ns: entradas = ns[-1]
                            elif "SALIDAS DE EFECTIVO" in l:
                                ns = extraer_numeros(l)
                                if ns: salidas = ns[-1]
                            elif "VALOR DEL PORTAFOLIO" in l and "TOTAL" not in l:
                                ns = extraer_numeros(l)
                                if len(ns) >= 2: valor_total = ns[1]
                                elif ns: valor_total = ns[0]
                        except Exception:
                            continue

                    smart = es_smart_cash(texto_p1)
                    portafolio = deuda = movimientos = []
                    movimientos_efectivo = []
                    try: deuda = extraer_deuda_gbm(pdf)
                    except Exception: pass
                    if smart:
                        try: movimientos_efectivo = extraer_movimientos_efectivo_smart_cash(pdf)
                        except Exception: pass
                    else:
                        try: portafolio = extraer_portafolio_gbm(pdf)
                        except Exception: pass
                        try: movimientos = extraer_movimientos_acciones(pdf)
                        except Exception: pass

                    tipo = "smart_cash" if smart else "gbm"
                    log.append(f"  -> Tipo: {'Smart Cash' if smart else 'GBM Regular'}")

                    clientes[nombre][tipo] = {
                        "resumen": pd.DataFrame([{
                            "Plataforma": "GBM Smart Cash" if smart else "GBM",
                            "Saldo Anterior": saldo_ant,
                            "Entradas de Efectivo": entradas,
                            "Salidas de Efectivo": salidas,
                            "Valor Total del Portafolio": valor_total,
                        }]),
                        "entradas": entradas, "salidas": salidas,
                        "valor_total": valor_total, "saldo_anterior": saldo_ant,
                        "portafolio": portafolio, "deuda": deuda,
                        "movimientos": movimientos,
                        "movimientos_efectivo": movimientos_efectivo,
                    }
        except Exception as e:
            log.append(f"  ERROR: {e}")

    return clientes, log


# ════════════════════════════════════════════════════════════
# GENERAR EXCEL DE EXTRACCION
# ════════════════════════════════════════════════════════════
def generar_excel_extraccion(clientes):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for nombre_cliente, cuentas in sorted(clientes.items()):
            hoja = nombre_cliente[:31]
            pd.DataFrame().to_excel(writer, sheet_name=hoja, index=False)
            fila = 0

            if cuentas["gbm"]:
                gbm = cuentas["gbm"]
                fila = escribir_seccion(writer, hoja, fila, "RESUMEN GBM",
                                        gbm["resumen"])
                if gbm.get("portafolio"):
                    df = pd.DataFrame(gbm["portafolio"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "PORTAFOLIO (ACCIONES / FIBRAS)", df)
                if gbm.get("deuda"):
                    df = pd.DataFrame(gbm["deuda"])
                    fila = escribir_seccion(writer, hoja, fila, "DEUDA", df)
                if gbm.get("movimientos"):
                    df_mov = pd.DataFrame(gbm["movimientos"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "COMPRA / VENTA DE ACCIONES", df_mov)
                    compras = df_mov[df_mov["Operacion"] == "Compra"]
                    ventas = df_mov[df_mov["Operacion"] == "Venta"]
                    fila = escribir_fila_total(
                        writer, hoja, fila, "COMPRA TOTAL",
                        compras["Neto"].sum() if len(compras) > 0 else 0.0)
                    fila = escribir_fila_total(
                        writer, hoja, fila, "VENTA TOTAL",
                        ventas["Neto"].sum() if len(ventas) > 0 else 0.0)
                    fila += 1

            if cuentas["smart_cash"]:
                sc = cuentas["smart_cash"]
                fila = escribir_seccion(writer, hoja, fila, "RESUMEN SMART CASH",
                                        sc["resumen"])
                if sc.get("deuda"):
                    df = pd.DataFrame(sc["deuda"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "DEUDA SMART CASH", df)
                if sc.get("movimientos_efectivo"):
                    df_efec = pd.DataFrame(sc["movimientos_efectivo"])
                    fila = escribir_seccion(writer, hoja, fila,
                                            "MOVIMIENTOS EFECTIVO SMART CASH",
                                            df_efec)
                    dep = df_efec[df_efec["Operacion"] == "Deposito"]
                    ret = df_efec[df_efec["Operacion"] == "Retiro"]
                    fila = escribir_fila_total(
                        writer, hoja, fila, "TOTAL DEPOSITOS",
                        dep["Monto"].sum() if len(dep) > 0 else 0.0)
                    fila = escribir_fila_total(
                        writer, hoja, fila, "TOTAL RETIROS",
                        ret["Monto"].sum() if len(ret) > 0 else 0.0)
                    fila += 1

            if cuentas["prestadero"]:
                fila = escribir_seccion(writer, hoja, fila, "RESUMEN PRESTADERO",
                                        cuentas["prestadero"]["resumen"])

            ajustar_columnas(writer.sheets[hoja])

    output.seek(0)
    return output


# ════════════════════════════════════════════════════════════
# GENERAR PDF DESDE EXCEL (tabla formateada)
# ════════════════════════════════════════════════════════════
def generar_pdf_desde_excel(excel_bytes, nombre_base):
    """Genera un PDF con las tablas del Excel usando solo librerias puras."""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError:
        return None

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    excel_bytes.seek(0)
    wb = load_workbook(excel_bytes, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Titulo de hoja (nombre del cliente)
        elements.append(Paragraph(
            f"<b>{sheet_name}</b>",
            styles['Title']))
        elements.append(Spacer(1, 12))

        # Leer todas las filas con datos
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column, values_only=True):
            fila = []
            for cell in row:
                if cell is None:
                    fila.append("")
                elif isinstance(cell, float):
                    fila.append(f"${cell:,.2f}" if abs(cell) >= 1 else
                                f"{cell:.6f}")
                else:
                    fila.append(str(cell))
            # Saltar filas completamente vacias
            if any(c.strip() for c in fila):
                data.append(fila)

        if not data:
            continue

        # Calcular anchos proporcionales
        n_cols = len(data[0])
        page_width = landscape(letter)[0] - 1*inch
        col_width = page_width / max(n_cols, 1)
        col_widths = [col_width] * n_cols

        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Estilo de tabla
        style_cmds = [
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]

        # Colorear filas de titulo (las que tienen texto en col A y nada despues)
        for i, row in enumerate(data):
            txt = row[0].upper() if row[0] else ""
            if any(kw in txt for kw in ["RESUMEN", "PORTAFOLIO", "DEUDA",
                                         "COMPRA", "MOVIMIENTOS"]):
                style_cmds.append(
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor("#2F5496")))
                style_cmds.append(
                    ('TEXTCOLOR', (0, i), (-1, i), colors.white))
                style_cmds.append(
                    ('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            elif any(kw in txt for kw in ["TOTAL", "COMPRA TOTAL",
                                           "VENTA TOTAL"]):
                style_cmds.append(
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor("#D6E4F0")))
                style_cmds.append(
                    ('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)
        elements.append(Spacer(1, 24))

    if elements:
        doc.build(elements)
        output.seek(0)
        return output
    return None


# ════════════════════════════════════════════════════════════
# FUNCIONES DE CONSOLIDACION
# ════════════════════════════════════════════════════════════
MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
MESES_INV = {v: k for k, v in MESES.items()}

ALIASES = {
    "FIBRAPL 14": ["FIBRA PL 14", "FIBRA PL14", "FIBRAPL14"],
    "FIHO 12": ["FIHO12"], "FMTY 14": ["FMTY14"],
    "FUNO 11": ["FUNO11"], "FIBRAMQ 12": ["FIBRAMQ12"],
    "DAHANOS 13": ["DANHOS 13", "DANHOS13", "DAHANOS13"],
}
_ALIAS_MAP = {}
for canonical, alts in ALIASES.items():
    group = {canonical.upper().strip()} | {a.upper().strip() for a in alts}
    for name in group:
        _ALIAS_MAP[name] = group


def normalizar(nombre):
    if not nombre or nombre == "-": return ""
    n = str(nombre).upper().strip()
    n = re.sub(r"\n", " ", n)
    n = re.sub(r"\s+", " ", n)
    return n


def instrumentos_coinciden(np_, nm_):
    np_ = normalizar(np_)
    nm_ = normalizar(nm_)
    if np_ == nm_: return True
    grupo = _ALIAS_MAP.get(np_, {np_})
    return nm_ in grupo


def encontrar_fila(ws, texto, col=1, rango=(1, 50)):
    for r in range(rango[0], rango[1]):
        v = ws.cell(r, col).value
        if v and texto in str(v).upper():
            return r
    return None


def valor_numerico(v, default=0.0):
    return float(v) if isinstance(v, (int, float)) else default


def actualizar_celda(ws, row, col, value, forzar=False):
    """Escribe en una celda, resolviendo merged ranges.
    Si la celda es parte de un merge, escribe en la celda principal.
    Si aun asi falla (MergedCell read-only), desmerge primero."""
    from openpyxl.cell.cell import MergedCell
    target_row, target_col = row, col
    merge_range = None
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            target_row, target_col = rng.min_row, rng.min_col
            merge_range = rng
            break
    celda = ws.cell(target_row, target_col)
    # Si sigue siendo MergedCell, desmerge el rango y reintenta
    if isinstance(celda, MergedCell) and merge_range:
        try:
            ws.unmerge_cells(str(merge_range))
        except Exception:
            pass
        celda = ws.cell(target_row, target_col)
    if not forzar and isinstance(celda.value, str) and str(celda.value).startswith("="):
        return
    try:
        celda.value = value
    except AttributeError:
        # Ultimo recurso: desmerge todo lo que toque esta celda
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    pass
        ws.cell(row, col).value = value


def copiar_formato_fila(ws, fila_origen, fila_destino):
    from openpyxl.cell.cell import MergedCell
    for col in range(1, 16):
        src = ws.cell(fila_origen, col)
        dst = ws.cell(fila_destino, col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        try:
            dst.font = copy(src.font); dst.fill = copy(src.fill)
            dst.border = copy(src.border); dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
        except AttributeError:
            continue


def insertar_instrumento(ws, fila_insercion, datos, fila_ref, periodo):
    """Inserta una nueva fibra/accion ANTES de la fila indicada."""
    ws.insert_rows(fila_insercion)
    nueva_fila = fila_insercion
    copiar_formato_fila(ws, fila_ref, nueva_fila)

    emisora = datos["emisora"]
    valor = datos["valor_a_mercado"]
    compra = datos["compra_neto"]
    venta = datos["venta_neto"]

    # B = inversion original (lo que se compro)
    b = compra if compra > 0 else valor
    c = valor
    e = c - b
    f = (e / b) if b > 0 else 0.0
    g = c + venta - compra  # ganancia del mes (instrumento nuevo, old_c=0)
    h = (g / b) if b > 0 else 0.0

    ws.cell(nueva_fila, 1).value = emisora            # A - Instrumento
    ws.cell(nueva_fila, 2).value = round(b, 2)        # B - Saldo Inicial
    ws.cell(nueva_fila, 3).value = round(c, 2)        # C - Saldo Total
    if periodo:
        ws.cell(nueva_fila, 4).value = (
            f"{periodo['mes_nombre'].lower()} {periodo['anio']}")  # D
    ws.cell(nueva_fila, 5).value = round(e, 2)        # E - Ganancia Hist
    ws.cell(nueva_fila, 6).value = round(f, 10)       # F - Ganancia Hist %
    ws.cell(nueva_fila, 7).value = round(g, 2)        # G - Gan/Perd Mes
    ws.cell(nueva_fila, 8).value = round(h, 10)       # H - Gan/Perd Mes %
    ws.cell(nueva_fila, 9).value = "RENTA VARIABLE"   # I - Clasificacion
    ws.cell(nueva_fila, 10).value = round(venta, 2)   # J - Retiros
    ws.cell(nueva_fila, 11).value = round(compra, 2)  # K - Depositos
    ws.cell(nueva_fila, 14).value = round(c, 2)       # N - Total
    ws.cell(nueva_fila, 15).value = "GBM"             # O - Proveedor

    return fila_insercion + 1


def expandir_formulas_totales(ws, fila_totales):
    patron = re.compile(r"(SUM\([A-Z]+)(\d+)(:[A-Z]+)(\d+)(\))")
    nueva_fin = fila_totales - 1
    for col in range(2, 16):
        celda = ws.cell(fila_totales, col)
        val = celda.value
        if not isinstance(val, str) or not val.startswith("="):
            continue
        nueva = patron.sub(
            lambda m: (f"{m.group(1)}{m.group(2)}"
                       f"{m.group(3)}{nueva_fin}{m.group(5)}"), val)
        if nueva != val:
            celda.value = nueva


def leer_instrumentos_master(ws, fila_header, fila_totales):
    instrumentos = []
    r = fila_header + 1
    while r < fila_totales:
        nombre = ws.cell(r, 1).value
        if nombre and str(nombre).strip() and str(nombre).strip() != "-":
            fila_fin = r
            for rng in ws.merged_cells.ranges:
                if rng.min_row == r and rng.min_col == 1:
                    fila_fin = rng.max_row; break
            instrumentos.append({
                "fila": r, "fila_fin": fila_fin,
                "nombre": str(nombre).strip(),
                "B": ws.cell(r, 2).value, "C": ws.cell(r, 3).value,
                "D": ws.cell(r, 4).value, "E": ws.cell(r, 5).value,
                "F": ws.cell(r, 6).value, "G": ws.cell(r, 7).value,
                "H": ws.cell(r, 8).value, "I": ws.cell(r, 9).value,
                "J": ws.cell(r, 10).value, "K": ws.cell(r, 11).value,
                "L": ws.cell(r, 12).value, "M": ws.cell(r, 13).value,
                "N": ws.cell(r, 14).value, "O": ws.cell(r, 15).value,
            })
            r = fila_fin + 1
        else:
            r += 1
    return instrumentos


def extraer_periodo_pdf_from_uploaded(pdf, plataforma):
    texto = pdf.pages[0].extract_text() or ""
    if plataforma == "Prestadero":
        m = re.search(
            r"(\d{4}).(\d{2}).(\d{2})\s+al\s+(\d{4}).(\d{2}).(\d{2})", texto)
        if m:
            anio, mes = int(m.group(4)), int(m.group(5))
            dia_fin = int(m.group(6))
            nombre_mes = MESES_INV.get(mes, str(mes))
            return {"mes": mes, "anio": anio, "mes_nombre": nombre_mes,
                    "periodo": f"01-{dia_fin} {nombre_mes[:3]} DE {anio}"}
    else:
        m = re.search(
            r"DEL\s+(\d+)\s+AL\s+(\d+)\s+DE\s+(\w+)\s+DE\s+(\d{4})",
            texto, re.I)
        if m:
            dia_ini, dia_fin = int(m.group(1)), int(m.group(2))
            nombre_mes = m.group(3).upper()
            anio = int(m.group(4))
            mes = MESES.get(nombre_mes, 0)
            return {"mes": mes, "anio": anio, "mes_nombre": nombre_mes,
                    "periodo": (f"{dia_ini:02d}-{dia_fin} "
                                f"{nombre_mes[:3]} DE {anio}")}
    return None


def _mejor_match_deuda(old_c, fuentes):
    if not fuentes: return None, None
    if old_c <= 0:
        k = min(fuentes, key=lambda k: fuentes[k]["valor"])
        return k, fuentes[k]
    mejor_key = min(fuentes, key=lambda k: abs(fuentes[k]["valor"] - old_c))
    return mejor_key, fuentes[mejor_key]


def buscar_hoja(wb, nombre_cliente):
    nc = nombre_cliente.upper().strip()
    for s in wb.sheetnames:
        if s.upper().strip() == nc: return s
    for s in wb.sheetnames:
        if nc in s.upper() or s.upper().strip() in nc: return s
    partes = nc.split()
    for s in wb.sheetnames:
        su = s.upper()
        if sum(1 for p in partes if p in su) >= 2: return s
    return None


def actualizar_hoja(ws, datos, nombre_hoja, log):
    """Actualiza una hoja del maestro con datos de PDFs.

    REGLAS DE SALDO INICIAL (B):
    - VENTA de fibra/accion  -> B = B_anterior - monto_venta (resta)
    - COMPRA de fibra/accion -> B = B_anterior + monto_compra (suma)
    - DEUDA y EFECTIVO       -> B no cambia
    - Fibras nuevas se insertan ANTES de EFECTIVO GBM
    """
    gbm = datos.get("gbm")
    smart_cash = datos.get("smart_cash")
    prestadero = datos.get("prestadero")
    periodo = datos.get("periodo")

    fila_header = encontrar_fila(ws, "INSTRUMENTO") or 23
    fila_totales = encontrar_fila(ws, "TOTALES",
                                   rango=(fila_header, fila_header + 40))
    if not fila_totales:
        log.append(f"  ⚠ No se encontro TOTALES en '{nombre_hoja}'")
        return

    instrumentos = leer_instrumentos_master(ws, fila_header, fila_totales)
    log.append(f"  Instrumentos encontrados: {len(instrumentos)}")

    # Preparar datos del PDF
    pdf_port = {}
    compras_map = defaultdict(float)
    ventas_map = defaultdict(float)

    if gbm:
        for item in gbm.get("portafolio", []):
            key = normalizar(item["Emisora"])
            pdf_port[key] = item["Valor a Mercado"]
        for mov in gbm.get("movimientos", []):
            key = normalizar(mov["Emisora"])
            if mov["Operacion"] == "Compra":
                compras_map[key] += mov["Neto"]
            else:
                ventas_map[key] += mov["Neto"]

    deuda_gbm_total = (sum(d["Valor del Reporto"]
                           for d in gbm.get("deuda", [])) if gbm else 0.0)
    deuda_sc_total = (sum(d["Valor del Reporto"]
                          for d in smart_cash.get("deuda", []))
                      if smart_cash else 0.0)
    sc_entradas = smart_cash.get("entradas", 0) if smart_cash else 0
    sc_salidas = smart_cash.get("salidas", 0) if smart_cash else 0

    # Pool de fuentes DEUDA
    fuentes_deuda = {}
    if prestadero:
        fuentes_deuda["prestadero"] = {
            "valor": prestadero["valor"], "retiros": prestadero["retiros"],
            "depositos": prestadero["abonos"],
            "interes": prestadero["interes"], "tipo": "prestadero",
        }
    if smart_cash and deuda_sc_total > 0:
        fuentes_deuda["smart_cash"] = {
            "valor": deuda_sc_total, "retiros": sc_salidas,
            "depositos": sc_entradas, "interes": 0.0, "tipo": "smart_cash",
        }
    if gbm and deuda_gbm_total > 0:
        fuentes_deuda["gbm_deuda"] = {
            "valor": deuda_gbm_total, "retiros": 0.0,
            "depositos": 0.0, "interes": 0.0, "tipo": "gbm_deuda",
        }

    matched_pdf_keys = set()
    deuda_gbm_matched = 0.0
    efectivo_instr = None

    for instr in instrumentos:
        fila = instr["fila"]
        nom = instr["nombre"]
        nom_n = normalizar(nom)
        old_c = valor_numerico(instr["C"])
        old_b = valor_numerico(instr["B"])
        matched = False
        new_c = new_g = None
        new_j = new_k = 0.0
        es_deuda = "DEUDA" in nom_n

        # ── EFECTIVO GBM: diferir al final ──
        if "EFECTIVO" in nom_n and "GBM" in nom_n:
            efectivo_instr = instr; continue

        # ── DEUDA con proveedor ──
        elif es_deuda and instr["O"]:
            prov = normalizar(str(instr["O"]))
            fuente_key = None
            if "PRESTADERO" in prov and "prestadero" in fuentes_deuda:
                fuente_key = "prestadero"
            elif "SMART" in prov and "smart_cash" in fuentes_deuda:
                fuente_key = "smart_cash"
            elif "GBM" in prov and "gbm_deuda" in fuentes_deuda:
                fuente_key = "gbm_deuda"
            if fuente_key:
                mf = fuentes_deuda[fuente_key]
                new_c = mf["valor"]
                new_j = mf["retiros"]
                new_k = mf["depositos"]
                new_g = (mf["interes"] if mf["tipo"] == "prestadero"
                         else new_c - old_c)
                matched = True
                if mf["tipo"] == "gbm_deuda":
                    deuda_gbm_matched += new_c
                log.append(
                    f"  DEUDA (fila {fila}) <- {fuente_key.upper()}"
                    f" ${new_c:,.2f}")
                del fuentes_deuda[fuente_key]

        # ── DEUDA sin proveedor ──
        elif es_deuda and not instr["O"] and fuentes_deuda:
            mk, mf = _mejor_match_deuda(old_c, fuentes_deuda)
            if mf:
                new_c = mf["valor"]
                new_j = mf["retiros"]
                new_k = mf["depositos"]
                new_g = (mf["interes"] if mf["tipo"] == "prestadero"
                         else new_c - old_c)
                matched = True
                if mf["tipo"] == "gbm_deuda":
                    deuda_gbm_matched += new_c
                del fuentes_deuda[mk]

        # ── FIBRAS / ACCIONES por nombre ──
        else:
            for pdf_key, pdf_valor in pdf_port.items():
                if instrumentos_coinciden(pdf_key, nom_n):
                    new_c = pdf_valor
                    compra_neto = compras_map.get(pdf_key, 0)
                    venta_neto = ventas_map.get(pdf_key, 0)
                    new_j = venta_neto   # Retiros = ventas
                    new_k = compra_neto  # Depositos = compras
                    # G = ganancia organica del mes
                    new_g = new_c - old_c + new_j - new_k
                    matched = True
                    matched_pdf_keys.add(pdf_key)
                    break

        if not matched:
            continue

        # ══════════════════════════════════════════
        # ACTUALIZAR SALDO INICIAL (B)
        # Venta -> resta del saldo inicial
        # Compra -> suma al saldo inicial
        # Deuda/Efectivo -> B no cambia
        # ══════════════════════════════════════════
        if not es_deuda:
            # B = B_anterior + compras - ventas
            new_b = old_b + new_k - new_j
            actualizar_celda(ws, fila, 2, round(new_b, 2))
            old_b = new_b

        # C = Saldo Total actual (valor a mercado del PDF)
        actualizar_celda(ws, fila, 3, round(new_c, 2))

        # E = Ganancia Historica = C - B
        new_e = new_c - old_b
        actualizar_celda(ws, fila, 5, round(new_e, 2))

        # F = Ganancia Historica % = E / B
        new_f = (new_e / old_b) if old_b > 0 else 0.0
        actualizar_celda(ws, fila, 6, round(new_f, 10))

        # G = Ganancia / Perdida del Mes
        if new_g is not None:
            actualizar_celda(ws, fila, 7, round(new_g, 2))

        # H = Ganancia / Perdida del Mes %
        if new_g is not None and old_b > 0:
            actualizar_celda(ws, fila, 8, round(new_g / old_b, 10))

        # J = Retiros,  K = Depositos
        actualizar_celda(ws, fila, 10, round(new_j, 2))
        actualizar_celda(ws, fila, 11, round(new_k, 2))

        # N = Total (= C)
        actualizar_celda(ws, fila, 14, round(new_c, 2))

        log.append(f"  OK {nom:<20s} C=${new_c:>12,.2f}"
                   f"  B=${old_b:>12,.2f}")

    # ── EFECTIVO GBM (siempre al final) ──
    if efectivo_instr and gbm:
        fila = efectivo_instr["fila"]
        old_c = valor_numerico(efectivo_instr["C"])
        new_c = round(
            gbm["valor_total"] - sum(pdf_port.values()) - deuda_gbm_matched, 2)
        if new_c < 0:
            new_c = 0.0
        new_g = new_c - old_c
        actualizar_celda(ws, fila, 3, round(new_c, 2))
        actualizar_celda(ws, fila, 5, "-")
        actualizar_celda(ws, fila, 6, "-")
        actualizar_celda(ws, fila, 7, round(new_g, 2))
        actualizar_celda(ws, fila, 8, "-")
        actualizar_celda(ws, fila, 10, 0.0)
        actualizar_celda(ws, fila, 11, 0.0)
        actualizar_celda(ws, fila, 14, round(new_c, 2))
        log.append(f"  OK {'EFECTIVO GBM':<20s} C=${new_c:>12,.2f}")

    # ── INSERTAR NUEVAS FIBRAS/ACCIONES (antes de EFECTIVO GBM) ──
    nuevos = []
    for pk, pv in pdf_port.items():
        if pk not in matched_pdf_keys:
            c = compras_map.get(pk, 0)
            v = ventas_map.get(pk, 0)
            if pv == 0 and c == 0 and v == 0:
                continue
            nuevos.append({
                "emisora": pk, "valor_a_mercado": pv,
                "compra_neto": c, "venta_neto": v,
            })

    if nuevos:
        # Insertar ANTES de EFECTIVO GBM para que quede al final
        if efectivo_instr:
            fi = efectivo_instr["fila"]
        else:
            fi = fila_totales
        fr = (instrumentos[-1]["fila"] if instrumentos
              else fila_header + 1)
        for n in nuevos:
            insertar_instrumento(ws, fi, n, fr, periodo)
            fi += 1
            fila_totales += 1
            log.append(
                f"  NUEVO: {n['emisora']:<16s}"
                f"  C=${n['valor_a_mercado']:>12,.2f}"
                f"  Compra=${n['compra_neto']:>10,.2f}")
        expandir_formulas_totales(ws, fila_totales)

    # Limpiar debajo de TOTALES
    ultima = ws.max_row
    if ultima > fila_totales:
        ws.delete_rows(fila_totales + 1, ultima - fila_totales)

    # Periodo
    if periodo:
        actualizar_celda(ws, 2, 9,
                         f"CORTE MENSUAL {periodo['mes_nombre']}",
                         forzar=True)
        actualizar_celda(ws, 3, 9, periodo["periodo"], forzar=True)


def consolidar_con_maestro(clientes, maestro_file, pdfs_files):
    """Consolida datos extraidos con el maestro."""
    log = []

    # Extraer periodos
    for archivo in pdfs_files:
        try:
            archivo.seek(0)
            with pdfplumber.open(archivo) as pdf:
                texto = pdf.pages[0].extract_text() or ""
                plataforma = detectar_plataforma(texto)
                nombre = extraer_nombre_cliente(pdf, plataforma)
                periodo = extraer_periodo_pdf_from_uploaded(pdf, plataforma)
                if nombre.upper() in {k.upper() for k in clientes}:
                    for k in clientes:
                        if k.upper() == nombre.upper():
                            if periodo and (
                                clientes[k].get("periodo") is None
                                or periodo["mes"] > clientes[k].get(
                                    "periodo", {}).get("mes", 0)
                            ):
                                clientes[k]["periodo"] = periodo
        except Exception:
            continue

    # Nombre de salida
    any_periodo = None
    for c in clientes.values():
        if c.get("periodo"):
            any_periodo = c["periodo"]; break
    mes_nombre = any_periodo["mes_nombre"] if any_periodo else "ACTUALIZADO"
    anio = any_periodo["anio"] if any_periodo else ""
    salida_nombre = f"ESTADOS DE CUENTA {mes_nombre} {anio}.xlsx".strip()

    # Cargar y actualizar maestro
    maestro_file.seek(0)
    wb = load_workbook(maestro_file)
    log.append(f"Hojas en maestro: {', '.join(wb.sheetnames)}")

    clientes_ok = 0
    for nombre_cliente, datos in sorted(clientes.items()):
        hoja = buscar_hoja(wb, nombre_cliente)
        if hoja:
            log.append(f"\n{nombre_cliente} -> Hoja: '{hoja}'")
            try:
                actualizar_hoja(wb[hoja], datos, hoja, log)
                clientes_ok += 1
            except Exception as e:
                log.append(f"  ERROR: {e}")
        else:
            log.append(f"\n⚠ {nombre_cliente} — Sin hoja en maestro")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log.append(f"\nClientes actualizados: {clientes_ok}")
    return output, salida_nombre, log


# ════════════════════════════════════════════════════════════
# INTERFAZ STREAMLIT — FLUJO PASO A PASO
# ════════════════════════════════════════════════════════════

# Header
st.markdown("""
<div class="main-header">
    <h1>📊 Automatizacion de Estados de Cuenta</h1>
    <p style="font-size: 1.1rem; color: #666;">
        Extrae datos de estados de cuenta GBM / Prestadero y actualiza
        el maestro anterior automaticamente.
    </p>
</div>
""", unsafe_allow_html=True)

st.divider()

# ────────────────────────────────────────────────────
# PASO 1: SUBIR PDFs DEL MES
# ────────────────────────────────────────────────────
st.markdown("""
<div class="step-header">
    📄 Paso 1 — Sube los estados de cuenta del mes (PDFs)
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="info-box">
    Selecciona todos los PDFs de estados de cuenta del mes actual
    (GBM Regular, Smart Cash y/o Prestadero).
    Se extraeran automaticamente los datos de cada uno.
</div>
""", unsafe_allow_html=True)
st.write("")

pdfs = st.file_uploader(
    "Arrastra o selecciona los PDFs del mes",
    type=["pdf"],
    accept_multiple_files=True,
    key="pdfs_mes",
    label_visibility="collapsed",
)

# Procesar PDFs automaticamente al subirlos
if pdfs:
    if ("clientes" not in st.session_state
            or st.session_state.get("_pdfs_hash") != len(pdfs)):
        with st.spinner("Extrayendo datos de los PDFs..."):
            clientes, log_ext = procesar_pdfs(pdfs)
            st.session_state["clientes"] = clientes
            st.session_state["log_ext"] = log_ext
            st.session_state["_pdfs_hash"] = len(pdfs)

    clientes = st.session_state["clientes"]
    log_ext = st.session_state["log_ext"]

    if clientes:
        st.success(
            f"Se detectaron **{len(clientes)} cliente(s)** "
            f"en **{len(pdfs)} PDF(s)**")

        # Mostrar resumen compacto
        cols_clientes = st.columns(min(len(clientes), 4))
        for i, (nombre, cuentas) in enumerate(sorted(clientes.items())):
            tipos = []
            if cuentas["gbm"]: tipos.append("GBM")
            if cuentas["smart_cash"]: tipos.append("Smart Cash")
            if cuentas["prestadero"]: tipos.append("Prestadero")
            with cols_clientes[i % len(cols_clientes)]:
                st.metric(label=nombre, value=", ".join(tipos))

        # Detalles expandibles
        with st.expander("Ver detalle de datos extraidos"):
            for nombre, cuentas in sorted(clientes.items()):
                st.subheader(nombre)
                if cuentas["gbm"]:
                    st.caption("GBM Regular")
                    st.dataframe(cuentas["gbm"]["resumen"],
                                 use_container_width=True, hide_index=True)
                    if cuentas["gbm"].get("portafolio"):
                        st.caption("Portafolio")
                        st.dataframe(
                            pd.DataFrame(cuentas["gbm"]["portafolio"]),
                            use_container_width=True, hide_index=True)
                    if cuentas["gbm"].get("movimientos"):
                        st.caption("Movimientos (Compras/Ventas)")
                        st.dataframe(
                            pd.DataFrame(cuentas["gbm"]["movimientos"]),
                            use_container_width=True, hide_index=True)
                if cuentas["smart_cash"]:
                    st.caption("Smart Cash")
                    st.dataframe(cuentas["smart_cash"]["resumen"],
                                 use_container_width=True, hide_index=True)
                if cuentas["prestadero"]:
                    st.caption("Prestadero")
                    st.dataframe(cuentas["prestadero"]["resumen"],
                                 use_container_width=True, hide_index=True)
                st.divider()

        with st.expander("Ver log de procesamiento"):
            for linea in log_ext:
                st.text(linea)

        # Descargar reporte de extraccion
        st.write("")
        st.markdown("**Descargar reporte de extraccion:**")
        excel_ext = generar_excel_extraccion(clientes)

        col_dl1, col_dl2, _ = st.columns([1, 1, 2])
        with col_dl1:
            st.download_button(
                label="⬇ Descargar Excel",
                data=excel_ext,
                file_name="Reporte_Extraccion.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
            )
        with col_dl2:
            excel_ext_copy = generar_excel_extraccion(clientes)
            pdf_ext = generar_pdf_desde_excel(excel_ext_copy,
                                               "Reporte_Extraccion")
            if pdf_ext:
                st.download_button(
                    label="⬇ Descargar PDF",
                    data=pdf_ext,
                    file_name="Reporte_Extraccion.pdf",
                    mime="application/pdf",
                )
            else:
                st.info("PDF no disponible (instalar reportlab)")
    else:
        st.warning("No se encontraron datos en los PDFs subidos.")

# ────────────────────────────────────────────────────
# PASO 2: SUBIR MAESTRO ANTERIOR
# ────────────────────────────────────────────────────
st.write("")
st.write("")
st.markdown("""
<div class="step-header">
    📁 Paso 2 — Sube el Excel maestro del mes anterior
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="info-box">
    Este es el archivo Excel del mes anterior. Se actualizaran
    automaticamente los saldos de cada instrumento:
    <b>compras se suman</b> al saldo inicial,
    <b>ventas se restan</b>.
    Las fibras nuevas se agregan antes del Efectivo GBM.
</div>
""", unsafe_allow_html=True)
st.write("")

maestro = st.file_uploader(
    "Arrastra o selecciona el Excel maestro del mes anterior (.xlsx)",
    type=["xlsx"],
    key="maestro_upload",
    label_visibility="collapsed",
)

# ────────────────────────────────────────────────────
# PASO 3: CONSOLIDAR
# ────────────────────────────────────────────────────
if pdfs and maestro and "clientes" in st.session_state:
    st.write("")
    st.write("")
    st.markdown("""
    <div class="step-header">
        🔄 Paso 3 — Generar estado de cuenta consolidado
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box">
        Se cruzaran los datos extraidos de los PDFs con el maestro anterior.
        Se actualizaran saldos, ganancias, compras, ventas y deuda
        de cada cliente.
    </div>
    """, unsafe_allow_html=True)
    st.write("")

    if st.button("🚀  Generar Estado Consolidado", type="primary",
                  use_container_width=True):

        clientes = st.session_state["clientes"]

        progress = st.progress(0, text="Consolidando datos...")
        progress.progress(30, text="Procesando clientes...")

        excel_cons, nombre_salida, log_cons = consolidar_con_maestro(
            clientes, maestro, pdfs)

        progress.progress(80, text="Generando archivos de descarga...")

        # Tambien generar reporte de extraccion
        for f in pdfs:
            f.seek(0)
        clientes2, _ = procesar_pdfs(pdfs)
        excel_ext2 = generar_excel_extraccion(clientes2)

        progress.progress(100, text="Completado")

        st.write("")
        st.markdown("""
        <div class="success-box">
            <h3>✅ Consolidacion completada</h3>
        </div>
        """, unsafe_allow_html=True)

        with st.expander("Ver log de consolidacion"):
            for linea in log_cons:
                st.text(linea)

        # ── DESCARGAS ──
        st.write("")
        st.markdown("### Descargar archivos generados")

        c1, c2 = st.columns(2)

        with c1:
            st.markdown("**Estado Consolidado** (maestro actualizado)")
            st.download_button(
                label="⬇ Excel - Estado Consolidado",
                data=excel_cons,
                file_name=nombre_salida,
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
                type="primary",
                use_container_width=True,
            )
            # PDF del consolidado
            excel_cons_copy = io.BytesIO(excel_cons.getvalue())
            pdf_cons = generar_pdf_desde_excel(excel_cons_copy,
                                                nombre_salida)
            if pdf_cons:
                st.download_button(
                    label="⬇ PDF - Estado Consolidado",
                    data=pdf_cons,
                    file_name=nombre_salida.replace(".xlsx", ".pdf"),
                    mime="application/pdf",
                    use_container_width=True,
                )

        with c2:
            st.markdown("**Reporte de Extraccion** (datos crudos)")
            st.download_button(
                label="⬇ Excel - Reporte Extraccion",
                data=excel_ext2,
                file_name="Reporte_Extraccion.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
                use_container_width=True,
            )
            excel_ext2_copy = generar_excel_extraccion(clientes2)
            pdf_ext2 = generar_pdf_desde_excel(excel_ext2_copy,
                                                "Reporte_Extraccion")
            if pdf_ext2:
                st.download_button(
                    label="⬇ PDF - Reporte Extraccion",
                    data=pdf_ext2,
                    file_name="Reporte_Extraccion.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

elif pdfs and not maestro:
    st.write("")
    st.info("👆 Sube el Excel maestro del mes anterior para continuar "
            "con la consolidacion.")

# Footer
st.write("")
st.write("")
st.divider()
st.caption("Automatizacion de Estados de Cuenta v2.0 — GBM / Prestadero")
